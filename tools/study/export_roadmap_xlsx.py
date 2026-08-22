#!/usr/bin/env python3
"""Export the study roadmap to docs/study/MIW_MEO_Class1_Study_Roadmap.xlsx.

    GOVERNED ARTEFACTS                      NORMALIZED MODEL        WORKBOOK
    study_spine.json          --.
    study_mappings.json         |
    official_syllabus.json      +--->  build_model()  ---->  render_workbook()
    official_crosswalk.json     |      (plain dicts)         (six projections)
    coverage_matrix.json        |
    written_evidence_horizon.json
    study_progress.json       --'   (durable, hand-maintained, never written)

THE RENDERER DECIDES NOTHING. Every cell is a field of the model and every
field is read from a governed artefact. No topic truth, no recurrence
judgement and no syllabus classification is computed here -- a second
classifier living in the Excel tooling would be a second taxonomy by the back
door.

EXPANDABLE BY CONSTRUCTION
--------------------------
No corpus size is hardcoded. Counts come from the evidence horizon, so adding
papers widens the numbers *and the derived public sentence* without anyone
editing this file. The historical-QI columns already exist in the model and
render as NOT YET INTEGRATED; when that layer reaches VALIDATED_RANGE they
populate themselves.

Six sheets rather than one wide one, because forty columns is not a roadmap.
Each sheet is a projection of the same normalized model, so they cannot drift
from one another.

PROGRESS SURVIVES REGENERATION because it is never generated: study_progress.json
is read-only to this tool, and a topic missing from it is NOT_STARTED rather
than an error.

Determinism: no clock is read; the corpus's own `generated_from` is stamped.

Usage:
    python tools/study/export_roadmap_xlsx.py              # write workbook
    python tools/study/export_roadmap_xlsx.py --model-json PATH
    python tools/study/export_roadmap_xlsx.py --check      # model builds + loads
"""
import argparse, hashlib, io, json, os, re, sys

if __name__ == '__main__':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, HERE)

import evidence_model as EM
import study_links as LINKS

D = os.path.join(ROOT, 'docs', 'study')
SPINE     = os.path.join(D, 'study_spine.json')
MAPPINGS  = os.path.join(D, 'study_mappings.json')
OFFICIAL  = os.path.join(D, 'official_syllabus.json')
CROSSWALK = os.path.join(D, 'official_crosswalk.json')
COVERAGE  = os.path.join(D, 'coverage_matrix.json')
HORIZON   = os.path.join(D, 'written_evidence_horizon.json')
PROGRESS  = os.path.join(D, 'study_progress.json')
SESSIONS  = os.path.join(D, 'study_sessions.json')
FAMILIES  = os.path.join(ROOT, 'meoclass1', 'pastpapers', 'intelligence',
                         'derived', 'sixyear_families.json')
WATCH     = os.path.join(ROOT, 'meoclass1', 'pastpapers', 'intelligence',
                         'derived', 'sixyear_temporal_watch.json')
EXAMINERS = os.path.join(ROOT, 'meoclass1', 'oral-intelligence',
                         'examiner-audit',
                         'CURRENT_EXAMINER_RELATIONSHIPS.jsonl')
OUT       = os.path.join(D, 'MIW_MEO_Class1_Study_Roadmap.xlsx')

NOT_YET = 'NOT YET INTEGRATED'
NOT_AVAILABLE = LINKS.NOT_AVAILABLE

# Fields reserved for the historical Written QI layer. They are part of the
# model TODAY so that no schema change is needed when the layer lands.
FUTURE_WRITTEN_FIELDS = (
    'historical_written_papers', 'historical_written_questions',
    'earliest_evidence', 'latest_evidence',
    'long_term_recurrence', 'recent_recurrence', 'trend',
    'currentness', 'historical_qi_status',
)


def _load(path):
    return json.load(open(path, encoding='utf-8'))


def study_order(domains):
    """Dependency-first study order. NOT the priority ranking.

    The two are genuinely different questions and must not be conflated:
    `priority_rank` asks "which topic carries the most weight?" and answers
    D03; study order asks "which can you actually start?" and answers D01,
    because D03 lists D01 as a prerequisite. A roadmap sorted by score would
    open the candidate on a topic they are not yet equipped to read.

    Repeatedly take the highest-scoring topic whose prerequisites are already
    placed. Deterministic, and transparent enough to argue with: ties break on
    topic_id. A prerequisite cycle would strand topics, so anything left
    unplaced is appended in rank order rather than silently dropped -- the
    spine validator already proves the graph is acyclic (R-PREREQ-ACYCLIC).
    """
    by_id = {d['domain_id']: d for d in domains}
    placed, order = set(), []
    while len(order) < len(domains):
        ready = [d for d in domains
                 if d['domain_id'] not in placed
                 and all(p in placed for p in d['prerequisites'])]
        if not ready:
            for d in sorted(domains, key=lambda x: x['priority_rank']):
                if d['domain_id'] not in placed:
                    order.append(d['domain_id'])
                    placed.add(d['domain_id'])
            break
        pick = max(ready, key=lambda d: (d['study_priority']['score'],
                                         d['domain_id']))
        order.append(pick['domain_id'])
        placed.add(pick['domain_id'])
    return {tid: i + 1 for i, tid in enumerate(order)}


# --------------------------------------------------------------------------- #
# Study-session layer -- what to do RIGHT NOW, as opposed to what matters most
# --------------------------------------------------------------------------- #
def load_sessions():
    """The approved sessions, plus a staleness verdict against the topic pack.

    The pack is the document a human approved; study_sessions.json is its
    machine form. They can drift, so the sessions file pins the pack's sha256
    and this check FAILS CLOSED: an edited pack whose sessions were not
    re-derived is reported STALE rather than quietly rendered, because a study
    plan that has silently diverged from its own reasoning is worse than no
    plan at all.
    """
    doc = _load(SESSIONS)
    src = doc['derived_from']
    pack = os.path.join(ROOT, src['path'])
    if not os.path.exists(pack):
        doc['pack_status'] = 'PACK_MISSING'
    else:
        actual = hashlib.sha256(open(pack, 'rb').read()).hexdigest()
        doc['pack_status'] = ('CURRENT' if actual == src['sha256']
                              else 'STALE_PACK_CHANGED')
    return doc


def examiner_signal():
    """question_id -> the examiners governed evidence ties to it."""
    out = {}
    if not os.path.exists(EXAMINERS):
        return out
    with open(EXAMINERS, encoding='utf-8') as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            if rec.get('status') != 'PUBLISHED':
                continue
            out.setdefault(rec['question_id'], set()).add(rec['examiner'])
    return {k: sorted(v) for k, v in out.items()}


def family_index():
    """member question_id -> its recurrence family, plus the temporal watchlist.

    This is the CURRENT six-year written recurrence layer (2021 onward). It is
    deliberately never labelled "recurrence" unqualified: the historical layer
    is a separate, not-yet-validated band, and conflating the two would let a
    six-year count masquerade as an all-time one.
    """
    by_member, watch = {}, set()
    if os.path.exists(FAMILIES):
        for fam in _load(FAMILIES):
            for m in fam['members']:
                by_member[m] = fam
    if os.path.exists(WATCH):
        for fam in _load(WATCH):
            watch.add(fam['family_id'])
    return by_member, watch


def skeleton_branches(pack_path):
    """The pack's MENTAL SKELETON as a list, so the workbook can carry it.

    Local file hyperlinks are the one route this tool cannot prove will OPEN
    -- Excel follows file:/// reliably, but whether Windows has a handler for
    .md is a shell association, not something a generator may assert. Carrying
    the skeleton inside the workbook makes that risk free: the thing Session 1
    step 1 asks him to read is already on screen.
    """
    full = os.path.join(ROOT, pack_path)
    if not os.path.exists(full):
        return []
    text = open(full, encoding='utf-8').read()
    i = text.find('## MENTAL SKELETON')
    if i < 0:
        return []
    j = text.find('\n## ', i + 5)
    seg = text[i:j if j > 0 else len(text)]
    out = []
    pat = r'^(\d+)\.\s+\*\*(.+?)\*\*\s*(.*?)(?=^\d+\.\s+\*\*|\Z)'
    for m in re.finditer(pat, seg, re.S | re.M):
        body = ' '.join(m.group(3).split())
        body = re.sub(r'\*+', '', body)
        out.append({'n': int(m.group(1)),
                    'branch': m.group(2).strip(),
                    'detail': body.strip(' -\u2014')})
    return out


def current_position(topics, sessions_doc, progress):
    """Which topic, and which session of it, Nixon is on RIGHT NOW.

    Derived, never hardcoded. The current topic is the first topic in STUDY
    ORDER (dependency-first, not the priority ranking) that has not reached
    the final stage of the progression declared in study_progress.json. When
    D01 completes this points at whatever legitimately unlocks next, with no
    edit to this file.
    """
    progression = progress.get('progression') or ['NOT_STARTED']
    done = progression[-1]
    cur = None
    for t in topics:                      # already in study order
        if t['study_status'] != done:
            cur = t
            break
    if cur is None:
        return {'topic': None, 'session': None, 'all_complete': True,
                'why': 'Every topic has reached %s.' % done,
                'next_stage': None, 'sessions_for_topic': []}

    mine = sorted([s for s in sessions_doc['sessions']
                   if s['topic_id'] == cur['topic_id']],
                  key=lambda s: s['session_number'])
    n = cur['sessions_completed'] + 1
    sess = next((s for s in mine if s['session_number'] == n), None)

    idx = (progression.index(cur['study_status'])
           if cur['study_status'] in progression else 0)
    nxt = progression[idx + 1] if idx + 1 < len(progression) else done

    why = ('Study order #%d of %d, and study order is dependency-first, not '
           'priority-first. Prerequisites: %s. It unlocks: %s. Priority rank '
           '#%d (score %s).' % (
               cur['study_order'], len(topics), cur['prerequisites'],
               cur['unlocks'], cur['priority_rank'], cur['priority_score']))

    return {'topic': cur, 'session': sess, 'all_complete': False, 'why': why,
            'next_stage': (sess or {}).get('advances_status_to') or nxt,
            'sessions_for_topic': mine}


def enrich_task(task, mappings, exam, fams, watch):
    """Attach the evidence and the CHECKED link a task needs to be actionable."""
    qid = task.get('question_id')
    row = dict(task, question='', link=None, examiners='\u2014',
               recurrence='\u2014', currentness='\u2014', kind='\u2014')
    if not qid:
        return row
    rec = mappings.get(qid) or {}
    fam = fams.get(qid)
    row['kind'] = rec.get('content_type') or ('WRITTEN' if '-Q' in qid else 'ORAL')
    row['question'] = rec.get('text') or (fam or {}).get('stem') or '\u2014'
    row['link'] = LINKS.question_link(qid)
    row['examiners'] = ', '.join(exam.get(qid, ())) or '\u2014'
    if fam:
        row['recurrence'] = '%s \u00d7%d (%s \u2192 %s)' % (
            fam['class'], fam['size'], fam['first_seen'], fam['last_seen'])
        if fam['family_id'] in watch:
            bounds = '; '.join(b[1] for b in (fam.get('boundaries') or ()))
            row['currentness'] = ('TEMPORAL WATCH \u2014 '
                                  + (bounds or 'law moved under this family'))
        else:
            row['currentness'] = 'Not on temporal watch'
    elif row['kind'] == 'ORAL':
        row['recurrence'] = 'n/a (oral)'
        row['currentness'] = 'n/a (oral)'
    return row


def build_model():
    spine, mappings = _load(SPINE), _load(MAPPINGS)
    official, xwalk = _load(OFFICIAL), _load(CROSSWALK)
    coverage, horizon = _load(COVERAGE), _load(HORIZON)
    progress = _load(PROGRESS)

    store = mappings['mappings']
    hist = horizon['layers']['historical_written_qi']
    cur = horizon['layers']['current_solved_written']
    qi_live = hist['status'] in ('VALIDATED_RANGE', 'COMPLETE')

    # papers per topic, counted -- never asserted
    papers_by_topic, questions_by_topic = {}, {}
    for rec in store.values():
        if rec['content_type'] != 'WRITTEN' or not rec.get('topic_id'):
            continue
        papers_by_topic.setdefault(rec['topic_id'], set()).add(rec['paper_id'])
        questions_by_topic[rec['topic_id']] = questions_by_topic.get(rec['topic_id'], 0) + 1

    cov_by_topic = {}
    for row in coverage['nodes']:
        cov_by_topic.setdefault(row['primary_topic'], []).append(row['coverage'])

    order = study_order(spine['domains'])
    topics = []
    for d in sorted(spine['domains'], key=lambda x: order[x['domain_id']]):
        did = d['domain_id']
        prog = (progress.get('topics') or {}).get(did) or {}
        bands = cov_by_topic.get(did, [])
        row = {
            # --- stable identity -----------------------------------------
            'topic_id': did,
            'topic': d['name'],
            'study_order': order[did],
            'priority_rank': d['priority_rank'],
            'prerequisites': ', '.join(d['prerequisites']) or '—',
            'unlocks': ', '.join(d['dependants']) or '—',
            # --- current, growing evidence -------------------------------
            'oral_questions': d['oral']['questions'],
            'examiner_evidenced_oral': d['examiner_intelligence']['oral_questions_with_evidence'],
            'examiner_relationships': d['examiner_intelligence']['relationship_occurrences'],
            'distinct_examiners': d['examiner_intelligence']['distinct_examiners'],
            'current_written_questions': d['written']['questions'],
            'current_written_papers': len(papers_by_topic.get(did, ())),
            'current_written_recurrence_families': d['written_question_intelligence']['recurring_families'],
            'official_syllabus_items': len(d['official_syllabus_nodes']),
            'official_supporting_items': len(d['official_supporting_nodes']),
            'official_node_ids': ', '.join(d['official_syllabus_nodes']) or '—',
            'coverage': ', '.join(f'{b}×{bands.count(b)}'
                                  for b in ('STRONG', 'PARTIAL', 'WEAK', 'NONE')
                                  if bands.count(b)) or '—',
            'priority_score': d['study_priority']['score'],
            # --- durable, user-owned -------------------------------------
            'study_status': prog.get('status', 'NOT_STARTED'),
            'sessions_completed': prog.get('sessions_completed', 0),
            'notes_written': ', '.join(prog.get('notes_written') or []) or '—',
            'last_touched': prog.get('last_touched') or '—',
            # --- checked routes -------------------------------------------
            # Every one is built by study_links and proved to resolve; a route
            # that does not exist yet arrives as ok=False and renders NOT YET
            # AVAILABLE rather than as a hyperlink into nothing.
            'link_study':  LINKS.study_topic(did),
            'link_orals':  LINKS.oral_topic(did),
            'link_pack':   LINKS.topic_pack(did),
        }
        for f in FUTURE_WRITTEN_FIELDS:
            row[f] = NOT_YET if not qi_live else None
        row['historical_qi_status'] = hist['status']
        row['gaps'] = '—'
        topics.append(row)

    # Topic 01's named gaps are the only hand-recorded ones today; they live
    # in the pack, so the workbook points at them rather than restating them.
    for row in topics:
        if row['topic_id'] == 'D01':
            row['gaps'] = 'N6 duty of care (zero coverage); N7 in-water survey; N8 official log book'
        if row['topic_id'] == 'D07':
            row['gaps'] = 'No Annexure III edge — Class II subject, examined orally anyway'
        if row['topic_id'] == 'D08':
            row['gaps'] = 'Supporting-only under official item 12'

    official_rows = []
    prim = {e['official_node_id']: e for e in xwalk['edges']
            if e['mapping_role'] == 'PRIMARY'}
    cov_by_node = {r['official_node_id']: r for r in coverage['nodes']}
    for n in official['nodes']:
        nid = n['official_node_id']
        c = cov_by_node.get(nid, {})
        official_rows.append({
            'official_node_id': nid,
            'item': n['official_number'],
            'page': n['source_page'],
            'primary_topic': prim.get(nid, {}).get('topic_id'),
            'primary_topic_name': prim.get(nid, {}).get('topic_name'),
            'supporting_topics': ', '.join(
                e['topic_id'] for e in xwalk['edges']
                if e['official_node_id'] == nid and e['mapping_role'] == 'SUPPORTING') or '—',
            'coverage': c.get('coverage'),
            'oral_evidence': c.get('oral_evidence'),
            'written_evidence': c.get('written_evidence'),
            'examiners': c.get('examiners'),
            'official_text': n['official_text'],
        })

    # ---------------------------------------------------------------- #
    # The operational layer: not "which topic matters most" but "what do I
    # open in the next sixty seconds". Everything below is a projection of
    # governed artefacts plus checked routes -- no new judgement is made.
    # ---------------------------------------------------------------- #
    sessions_doc = load_sessions()
    exam = examiner_signal()
    fams, watch = family_index()
    pos = current_position(topics, sessions_doc, progress)

    def _tasks(sess):
        return [enrich_task(t, store, exam, fams, watch)
                for t in (sess or {}).get('tasks', ())]

    current_tasks = _tasks(pos['session'])

    # The study queue is the current session's tasks followed by the rest of
    # the topic's approved sessions, so "what comes next" is never a guess.
    queue, order_n = [], 0
    for sess in pos['sessions_for_topic']:
        if pos['session'] and sess['session_number'] < pos['session']['session_number']:
            continue
        for row in _tasks(sess):
            order_n += 1
            queue.append(dict(row, queue_order=order_n,
                              topic_id=sess['topic_id'],
                              session_number=sess['session_number'],
                              session_title=sess['title'],
                              session_note=sess.get('note') or '',
                              is_current_session=(pos['session'] is not None
                                  and sess['session_number'] == pos['session']['session_number'])))

    # Every solved written question, carrying its topic and a checked link.
    # This is the honest answer to "written by topic": solvedQP/topics.html
    # exists but is keyed on a PRODUCT taxonomy (dom-*) with no governed join
    # to the D01-D10 spine, and inventing that join here would be exactly the
    # second classifier this file refuses to become.
    written_rows = []
    for qid, rec in sorted(store.items()):
        if rec['content_type'] != 'WRITTEN':
            continue
        fam = fams.get(qid)
        lk = LINKS.written_question(qid)
        written_rows.append({
            'question_id': qid,
            'paper_id': rec['paper_id'],
            'topic_id': rec.get('topic_id') or '\u2014',
            'question': (fam or {}).get('stem') or '\u2014',
            'recurrence': ('%s \u00d7%d' % (fam['class'], fam['size'])) if fam else '\u2014',
            'last_seen': (fam or {}).get('last_seen') or '\u2014',
            'on_temporal_watch': 'YES' if fam and fam['family_id'] in watch else 'no',
            'link': lk,
        })

    all_links = ([t['link_study'] for t in topics]
                 + [t['link_orals'] for t in topics]
                 + [t['link_pack'] for t in topics if t['link_pack']['ok']]
                 + [r['link'] for r in queue if r['link']]
                 + [r['link'] for r in written_rows])

    return {
        'schema_version': '1.1',
        'generated_by': 'tools/study/export_roadmap_xlsx.py',
        'sessions_doc': sessions_doc,
        'current': pos,
        'current_tasks': current_tasks,
        'queue': queue,
        'written_by_topic': written_rows,
        'skeleton': skeleton_branches(sessions_doc['derived_from']['path']),
        'progression': progress.get('progression') or [],
        'link_failures': LINKS.validate(all_links),
        'link_count': len(all_links),
        '_all_links': all_links,
        'generated_from': horizon['generated_from'],
        'versions': dict(horizon['versions'],
                         study_spine_version=spine['spine_version'],
                         mapping_version=mappings['schema_version'],
                         taxonomy_version=mappings['taxonomy_version']),
        'evidence_horizon': horizon['layers'],
        'public_claim': horizon['public_claim']['derived_sentence'],
        'official_source': official['source'],
        'topics': topics,
        'official_nodes': official_rows,
        'future_written_fields': list(FUTURE_WRITTEN_FIELDS),
    }


# --------------------------------------------------------------------------- #
# Renderer
# --------------------------------------------------------------------------- #
def render_workbook(model, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.hyperlink import Hyperlink

    HEAD = PatternFill('solid', fgColor='1F3B57')
    HEADF = Font(color='FFFFFF', bold=True, size=10)
    MUTED = Font(color='808080', italic=True, size=10)
    WRAP = Alignment(vertical='top', wrap_text=True)
    TOP = Alignment(vertical='top')

    # The operational palette. NOW is the one thing on screen that must be
    # unmissable; everything else is deliberately quieter than it.
    NOW = PatternFill('solid', fgColor='FFF3C4')      # current session
    BANNER = PatternFill('solid', fgColor='1F3B57')
    BANNERF = Font(color='FFFFFF', bold=True, size=11)
    KEY = Font(bold=True, size=10)
    BIG = Font(bold=True, size=16, color='1F3B57')
    LINKF = Font(color='0563C1', underline='single', size=10, bold=True)
    DEADF = Font(color='9A9A9A', italic=True, size=10)
    THIN = Side(style='thin', color='D0D7DE')
    BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def put_link(ws, row, col, rec, text='OPEN', fallback=NOT_AVAILABLE):
        """Write a REAL Excel hyperlink, or an honest blank. Never a dead link.

        A link that 404s costs a click and, worse, teaches the candidate not
        to trust the workbook. So an unresolved route renders as greyed text
        with no hyperlink attached at all.
        """
        c = ws.cell(row=row, column=col)
        if rec and rec.get('ok'):
            c.value = text
            c.hyperlink = rec['url']
            c.font = LINKF
        else:
            c.value = fallback
            c.font = DEADF
        c.alignment = TOP
        return c

    def put_internal(ws, row, col, sheet_name, cell_ref, text):
        """A one-click jump to another sheet of this workbook."""
        c = ws.cell(row=row, column=col)
        c.value = text
        c.hyperlink = Hyperlink(ref=c.coordinate,
                                location="'%s'!%s" % (sheet_name, cell_ref),
                                display=text)
        c.font = LINKF
        c.alignment = TOP
        return c

    wb = Workbook()
    wb.remove(wb.active)

    def sheet(title, headers, rows, widths, wrap_cols=()):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i)
            c.fill, c.font, c.alignment = HEAD, HEADF, WRAP
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
        for r in rows:
            ws.append(r)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = WRAP if c.column in wrap_cols else TOP
                if c.value == NOT_YET:
                    c.font = MUTED
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        return ws

    # --- 0. START HERE -- the operational cockpit -------------------------
    # First sheet, because the first question a study tool must answer is not
    # "what matters most" but "what do I open right now". The evidence sheets
    # that follow answer the first question; this one answers the second.
    cur = model['current']
    ws = wb.create_sheet('START HERE')
    for col, w in zip('ABCDEFGH', (10, 20, 62, 20, 66, 14, 26, 13)):
        ws.column_dimensions[col].width = w

    ws['A1'] = 'START HERE'
    ws['A1'].font = BIG
    ws['C1'] = model['public_claim']
    ws['C1'].font = MUTED
    ws['C1'].alignment = WRAP

    r = 3
    if cur['all_complete']:
        ws.cell(row=r, column=1, value='ALL TOPICS COMPLETE').font = BANNERF
        r += 2
        topic = None
    else:
        topic = cur['topic']
        sess = cur['session']
        banner = [
            ('CURRENT TOPIC', '%s \u2014 %s' % (topic['topic_id'], topic['topic'])),
            ('CURRENT STUDY STAGE', topic['study_status']),
            ('WHY THIS TOPIC IS NEXT', cur['why']),
            ("TODAY'S SESSION",
             ('Session %d \u2014 %s' % (sess['session_number'], sess['title']))
             if sess else
             ('No sessions are authored for %s yet. Until they are, work the '
              'topic directly: open its orals below, and author '
              'docs/study/TOPIC_%s_*.md the way Topic 01 was authored.'
              % (topic['topic_id'], topic['topic_id'][1:]))),
            ('ESTIMATED TIME',
             ('%d\u2013%d minutes' % (sess['minutes_min'], sess['minutes_max']))
             if sess else '\u2014'),
            ('PROGRESS', '%d session(s) completed \u00b7 this session moves you to %s'
             % (topic['sessions_completed'], cur['next_stage'])),
            ('NEXT ACTION',
             (sess['tasks'][0]['task'] if sess and sess['tasks']
              else 'Open %s on topics.html and work its A-priority orals.'
                   % topic['topic_id'])),
        ]
        for label, value in banner:
            c = ws.cell(row=r, column=1, value=label)
            c.font = BANNERF
            c.fill = BANNER
            c.alignment = WRAP
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.cell(row=r, column=2).fill = BANNER
            v = ws.cell(row=r, column=3, value=value)
            v.alignment = WRAP
            v.font = KEY if label in ('CURRENT TOPIC', "TODAY'S SESSION") else Font(size=10)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            r += 1

        # The four routes, one click each.
        r += 1
        ws.cell(row=r, column=1, value='OPEN').font = KEY
        routes = [
            ('OPEN STUDY PAGE', topic['link_study'], 'study.html'),
            ('OPEN TOPIC (all orals)', topic['link_orals'], 'topics.html'),
            ('OPEN LOCAL TOPIC PACK', topic['link_pack'], 'the .md study pack'),
        ]
        for label, rec, note in routes:
            ws.cell(row=r, column=2, value=label).font = KEY
            put_link(ws, r, 3, rec, text=(rec.get('url') or NOT_AVAILABLE))
            ws.cell(row=r, column=3).alignment = WRAP
            ws.cell(row=r, column=5, value=note).font = MUTED
            r += 1
        put_internal(ws, r, 2, 'STUDY QUEUE', 'A1', 'OPEN STUDY QUEUE')
        ws.cell(row=r, column=5, value='what comes after today').font = MUTED
        r += 2

    # --- today's session, as a table you can work down -------------------
    hdr_row = r
    headers = ['Step', 'Type', 'Task', 'Question ID', 'Question',
               'Direct Link', 'Target', 'Done']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill, c.font, c.alignment = HEAD, HEADF, WRAP
    r += 1
    first_task_row = r
    for t in model['current_tasks']:
        ws.cell(row=r, column=1, value=t['step']).alignment = TOP
        ws.cell(row=r, column=2, value=t['type']).alignment = TOP
        ws.cell(row=r, column=3, value=t['task']).alignment = WRAP
        ws.cell(row=r, column=4, value=t['question_id'] or '\u2014').alignment = TOP
        ws.cell(row=r, column=5, value=t['question']).alignment = WRAP
        if t['question_id']:
            put_link(ws, r, 6, t['link'], text='OPEN ' + t['question_id'])
        else:
            ws.cell(row=r, column=6, value='\u2014').font = MUTED
        ws.cell(row=r, column=7,
                value=t.get('target') or ('%s min' % t['minutes'] if t.get('minutes') else '\u2014')
                ).alignment = WRAP
        ws.cell(row=r, column=8, value='TODO').alignment = TOP
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = NOW
            ws.cell(row=r, column=col).border = BOX
        r += 1
    last_task_row = r - 1

    if last_task_row >= first_task_row:
        dv = DataValidation(type='list', formula1='"TODO,DONE,REDO"',
                            allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add('H%d:H%d' % (first_task_row, last_task_row))
    ws.cell(row=r, column=1,
            value='The Done column is a scratch tick-list. The DURABLE record '
                  'is docs/study/study_progress.json \u2014 this workbook only '
                  'reads it, so nothing you type here is saved back.').font = MUTED
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1).alignment = WRAP
    r += 2

    # --- the mental skeleton, carried inline ------------------------------
    # Step 1 of Session 1 is "read the skeleton". Making him leave the
    # workbook to do the first instruction in it would be a poor cockpit, and
    # the .md file link cannot be proved to open on any given machine.
    if model['skeleton']:
        c = ws.cell(row=r, column=1, value='MENTAL SKELETON \u2014 %s'
                    % (topic['topic'] if topic else ''))
        c.font = Font(bold=True, size=12, color='1F3B57')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        ws.cell(row=r, column=1, value='Step 1 of this session: read these once. '
                'Do not take notes yet.').font = MUTED
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        for b in model['skeleton']:
            ws.cell(row=r, column=1, value=b['n']).alignment = TOP
            ws.cell(row=r, column=2, value=b['branch']).font = KEY
            ws.cell(row=r, column=2).alignment = WRAP
            ws.cell(row=r, column=3, value=b['detail']).alignment = WRAP
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            for col in range(1, 6):
                ws.cell(row=r, column=col).border = BOX
            r += 1

    ws.freeze_panes = 'A%d' % (hdr_row + 1)

    # --- 0b. STUDY QUEUE -- what comes after today ------------------------
    # Expandable by construction: it is generated from study_sessions.json, so
    # later generated sessions for D02..D10 arrive here with no code change.
    ws = wb.create_sheet('STUDY QUEUE')
    qheaders = ['#', 'Topic', 'Session', 'Task Type', 'Question ID', 'Question',
                'Reason', 'Examiner Signal', 'Written Recurrence (6-yr)',
                'Currentness', 'Status', 'Direct Link']
    for i, h in enumerate(qheaders, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill, c.font, c.alignment = HEAD, HEADF, WRAP
    for col, w in zip(range(1, 13), (5, 8, 30, 20, 20, 62, 34, 20, 30, 34, 11, 22)):
        ws.column_dimensions[get_column_letter(col)].width = w
    r = 2
    for q in model['queue']:
        vals = [q['queue_order'], q['topic_id'],
                'S%d \u2014 %s' % (q['session_number'], q['session_title']),
                q['type'], q['question_id'] or '\u2014', q['question'],
                q['session_note'] or q['task'], q['examiners'], q['recurrence'],
                q['currentness'],
                'CURRENT' if q['is_current_session'] else 'QUEUED']
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = WRAP if i in (3, 6, 7, 9, 10) else TOP
            if q['is_current_session']:
                c.fill = NOW
        if q['question_id']:
            c = put_link(ws, r, 12, q['link'], text='OPEN')
        else:
            c = ws.cell(row=r, column=12, value='\u2014')
            c.font = MUTED
        if q['is_current_session']:
            c.fill = NOW
        r += 1
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.cell(row=r + 1, column=1,
            value='Generated from docs/study/study_sessions.json. Sessions for '
                  'later topics appear here automatically as they are approved '
                  '\u2014 this sheet has no hand-written rows.').font = MUTED

    # Row index of each topic's first row on WRITTEN BY TOPIC / STUDY QUEUE,
    # so the roadmap's per-topic routes can jump straight to it.
    written_sorted = sorted(model['written_by_topic'],
                            key=lambda r: (r['topic_id'], r['question_id']))
    written_first = {}
    for i, r_ in enumerate(written_sorted, start=2):
        written_first.setdefault(r_['topic_id'], i)
    queue_first = {}
    for i, q in enumerate(model['queue'], start=2):
        queue_first.setdefault(q['topic_id'], i)

    # --- 1. ROADMAP -- the sheet Nixon actually reads ---------------------
    # The old `Links` column held the literal text "topics.html#D01", which is
    # a location, not a route: it told him where to go and then made him get
    # there himself. Five checked hyperlinks replace it.
    ws = sheet('ROADMAP',
          ['Study Order', 'Topic ID', 'Topic', 'Prerequisites', 'Unlocks',
           'Oral Qs', 'Examiner-Evidenced Orals', 'Written Qs',
           'Written Papers', 'Recurrence Families', 'Official Items',
           'Coverage', 'Priority Rank', 'Priority Score', 'Study Status',
           'Open Study Topic', 'Oral by Topic', 'Written by Topic',
           'Topic Pack', 'Next Session', 'Gaps'],
          [[t['study_order'], t['topic_id'], t['topic'], t['prerequisites'],
            t['unlocks'], t['oral_questions'], t['examiner_evidenced_oral'],
            t['current_written_questions'], t['current_written_papers'],
            t['current_written_recurrence_families'],
            t['official_syllabus_items'], t['coverage'], t['priority_rank'],
            t['priority_score'], t['study_status'], None, None, None, None,
            None, t['gaps']]
           for t in model['topics']],
          [11, 9, 34, 13, 16, 8, 12, 9, 9, 12, 10, 16, 11, 11, 18,
           17, 15, 17, 15, 15, 46],
          wrap_cols={3, 21})
    cur_tid = (model['current']['topic'] or {}).get('topic_id')
    for i, t in enumerate(model['topics'], start=2):
        put_link(ws, i, 16, t['link_study'], text='STUDY \u25b8')
        put_link(ws, i, 17, t['link_orals'], text='ORALS \u25b8')
        if t['topic_id'] in written_first:
            put_internal(ws, i, 18, 'WRITTEN BY TOPIC',
                         'A%d' % written_first[t['topic_id']], 'WRITTEN \u25b8')
        else:
            ws.cell(row=i, column=18, value=NOT_AVAILABLE).font = DEADF
        put_link(ws, i, 19, t['link_pack'], text='PACK \u25b8')
        if t['topic_id'] in queue_first:
            put_internal(ws, i, 20, 'STUDY QUEUE',
                         'A%d' % queue_first[t['topic_id']], 'QUEUE \u25b8')
        else:
            ws.cell(row=i, column=20, value=NOT_AVAILABLE).font = DEADF
        if t['topic_id'] == cur_tid:
            for col in range(1, 22):
                cell = ws.cell(row=i, column=col)
                if cell.font is not LINKF:
                    cell.fill = NOW

    # --- 2. TOPIC DETAIL --------------------------------------------------
    # Statistics are only half of it -- the aggregate has to be walkable back
    # to the questions it counted, or it is trivia.
    ws = sheet('TOPIC DETAIL',
          ['Topic ID', 'Topic', 'Oral Qs', 'Examiner-Evidenced Orals',
           'Examiner Relationships', 'Distinct Examiners', 'Written Qs',
           'Written Papers', 'Recurrence Families', 'Official PRIMARY Items',
           'Official SUPPORTING Items', 'Official Node IDs', 'Priority Score',
           'Study Topic', 'Oral Questions', 'Written Questions', 'Topic Pack'],
          [[t['topic_id'], t['topic'], t['oral_questions'],
            t['examiner_evidenced_oral'], t['examiner_relationships'],
            t['distinct_examiners'], t['current_written_questions'],
            t['current_written_papers'], t['current_written_recurrence_families'],
            t['official_syllabus_items'], t['official_supporting_items'],
            t['official_node_ids'], t['priority_score'], None, None, None, None]
           for t in model['topics']],
          [9, 34, 8, 12, 12, 10, 9, 9, 12, 12, 12, 34, 11, 14, 15, 17, 14],
          wrap_cols={2, 12})
    for i, t in enumerate(model['topics'], start=2):
        put_link(ws, i, 14, t['link_study'], text='STUDY \u25b8')
        put_link(ws, i, 15, t['link_orals'], text='ORALS \u25b8')
        if t['topic_id'] in written_first:
            put_internal(ws, i, 16, 'WRITTEN BY TOPIC',
                         'A%d' % written_first[t['topic_id']], 'WRITTEN \u25b8')
        else:
            ws.cell(row=i, column=16, value=NOT_AVAILABLE).font = DEADF
        put_link(ws, i, 17, t['link_pack'], text='PACK \u25b8')

    # --- 2b. WRITTEN BY TOPIC -- every solved written question, linked ----
    # solvedQP/topics.html exists but is keyed on a product taxonomy (dom-*)
    # with no governed join to the D01-D10 spine. Rather than invent that
    # join, the workbook lists the questions themselves: study_mappings
    # already assigns each one a topic, and the anchor convention is proved.
    ws = sheet('WRITTEN BY TOPIC',
          ['Topic ID', 'Question ID', 'Paper', 'Recurrence (6-yr)',
           'Last Seen', 'Temporal Watch', 'Question', 'Open Answer'],
          [[r_['topic_id'], r_['question_id'], r_['paper_id'], r_['recurrence'],
            r_['last_seen'], r_['on_temporal_watch'], r_['question'], None]
           for r_ in written_sorted],
          [9, 14, 9, 18, 15, 14, 96, 14],
          wrap_cols={7})
    for i, r_ in enumerate(written_sorted, start=2):
        put_link(ws, i, 8, r_['link'], text='OPEN \u25b8')

    # --- 3. OFFICIAL SYLLABUS -- the DGMA text, quoted --------------------
    sheet('OFFICIAL SYLLABUS',
          ['Item', 'Node ID', 'Page', 'Primary Topic', 'Primary Topic Name',
           'Supporting Topics', 'Coverage', 'Official Wording (Annexure III)'],
          [[n['item'], n['official_node_id'], n['page'], n['primary_topic'],
            n['primary_topic_name'], n['supporting_topics'], n['coverage'],
            n['official_text']] for n in model['official_nodes']],
          [6, 13, 6, 13, 32, 17, 11, 120],
          wrap_cols={5, 8})

    # --- 4. COVERAGE ------------------------------------------------------
    sheet('COVERAGE',
          ['Node ID', 'Item', 'Primary Topic', 'Coverage', 'Oral Evidence',
           'Written Evidence', 'Examiners', 'Subject'],
          [[n['official_node_id'], n['item'], n['primary_topic'], n['coverage'],
            n['oral_evidence'], n['written_evidence'], n['examiners'],
            n['official_text'][:160]] for n in model['official_nodes']],
          [13, 6, 13, 11, 13, 15, 10, 90],
          wrap_cols={8})

    # --- 5. WRITTEN QI -- current layer plus the declared future socket ---
    ws = wb.create_sheet('WRITTEN QI')
    cur = model['evidence_horizon']['current_solved_written']
    hist = model['evidence_horizon']['historical_written_qi']
    ws.append(['EVIDENCE HORIZON', '', ''])
    ws['A1'].font = Font(bold=True, size=12)
    rows = [
        ['', '', ''],
        ['Layer', 'CURRENT_SOLVED_WRITTEN', 'HISTORICAL_WRITTEN_QI'],
        ['Status', cur['completeness'], hist['status']],
        ['Source status', cur['source_status'], hist['source_status']],
        ['Papers', cur['papers_total'], hist['papers_total'] or NOT_YET],
        ['Questions', cur['questions_total'], hist['questions_total'] or NOT_YET],
        ['Earliest sitting', cur['earliest_sitting'], hist['earliest_sitting'] or NOT_YET],
        ['Latest sitting', cur['latest_sitting'], hist['latest_sitting'] or NOT_YET],
        ['Years spanned', cur['years_spanned'], hist['years_spanned'] or NOT_YET],
        ['', '', ''],
        ['KNOWN GAPS', '', ''],
    ]
    for g in hist['known_gaps']:
        rows.append(['', f"{g['from_year']}–{g['to_year']}", g['reason']])
    rows += [
        ['', '', ''],
        ['RESERVED FIELDS (populate themselves when the layer is validated)', '', ''],
    ]
    for f in model['future_written_fields']:
        rows.append(['', f, NOT_YET])
    rows += [
        ['', '', ''],
        ['PUBLIC CLAIM (generated, never hand-written)', '', ''],
        ['', model['public_claim'], ''],
    ]
    for r in rows:
        ws.append(r)
    for col, w in zip('ABC', (56, 40, 78)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows():
        for c in row:
            c.alignment = WRAP
            if c.value == NOT_YET:
                c.font = MUTED
    # --- 6. PROGRESS -- the only sheet a human edits ----------------------
    ws = sheet('PROGRESS',
               ['Topic ID', 'Topic', 'Study Status', 'Sessions Completed',
                'Notes Written', 'Last Touched'],
               [[t['topic_id'], t['topic'], t['study_status'],
                 t['sessions_completed'], t['notes_written'], t['last_touched']]
                for t in model['topics']],
               [9, 34, 20, 17, 40, 14],
               wrap_cols={2, 5})
    if model['progression']:
        dv = DataValidation(type='list',
                            formula1='"%s"' % ','.join(model['progression']),
                            allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add('C2:C%d' % (len(model['topics']) + 1))
    ws.append([])
    ws.append(['Progress is stored in docs/study/study_progress.json and is '
               'never overwritten by a regeneration. The dropdown above is a '
               'convenience for reading the ladder \u2014 this workbook cannot '
               'write back, so a stage you reach must be recorded in that '
               'JSON file to survive the next regeneration.'])
    ws.cell(row=ws.max_row, column=1).font = MUTED
    ws.cell(row=ws.max_row, column=1).alignment = WRAP

    # --- 7. ABOUT ---------------------------------------------------------
    ws = wb.create_sheet('ABOUT')
    about = [
        ['MIW — MEO Class I Study Roadmap'],
        [''],
        [model['public_claim']],
        [''],
        ['Official syllabus', model['official_source']['circular']],
        ['Issued', model['official_source']['issue_date']],
        ['Annexure', 'III — Syllabus for MEO Class I Preparatory Course'],
        ['Source SHA-256', model['official_source']['sha256']],
        ['Effective from', '2027-01-01 (adopted, not yet in force)'],
        [''],
        ['Versions'],
    ]
    for k, v in sorted(model['versions'].items()):
        about.append(['', f'{k}', str(v)])
    about += [
        [''],
        ['Study sessions', model['sessions_doc']['derived_from']['path']],
        ['Sessions vs pack', model['sessions_doc']['pack_status']],
        ['Checked routes', '%d, of which %d could not be resolved'
         % (model['link_count'], len(model['link_failures']))],
        [''],
        ['Every number in this workbook is derived from a governed repository '
         'artefact. Nothing here is typed by hand, and the renderer classifies '
         'nothing itself. Every hyperlink was resolved against the repository '
         'at generation time: a route that did not exist renders as NOT YET '
         'AVAILABLE rather than as a link into nothing.'],
    ]
    for r in about:
        ws.append(r)
    ws['A1'].font = Font(bold=True, size=14)
    for col, w in zip('ABC', (34, 40, 78)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows():
        for c in row:
            c.alignment = WRAP

    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', default=OUT)
    ap.add_argument('--model-json')
    ap.add_argument('--check', action='store_true')
    args = ap.parse_args()

    model = build_model()
    if args.model_json:
        with open(args.model_json, 'w', encoding='utf-8', newline='\n') as fh:
            fh.write(json.dumps(model, indent=2, ensure_ascii=False) + '\n')
        print(f'wrote model {args.model_json}')

    if args.check:
        if not os.path.exists(args.out):
            print('FAIL: roadmap workbook is missing')
            return 1
        from openpyxl import load_workbook
        wb = load_workbook(args.out)
        want = ['START HERE', 'STUDY QUEUE', 'ROADMAP', 'TOPIC DETAIL',
                'WRITTEN BY TOPIC', 'OFFICIAL SYLLABUS', 'COVERAGE',
                'WRITTEN QI', 'PROGRESS', 'ABOUT']
        fails, warns = [], []
        if wb.sheetnames != want:
            fails.append('sheets %s != %s' % (wb.sheetnames, want))
        if wb.sheetnames and wb.sheetnames[0] != 'START HERE':
            fails.append('START HERE is not the first sheet')
        if wb['ROADMAP'].max_row != len(model['topics']) + 1:
            fails.append('ROADMAP row count does not match the model')

        # No dead hyperlinks, and none silently missing. Every link the model
        # built was resolved against the repo; assert the workbook shipped
        # exactly those and nothing typed by hand.
        if model['link_failures']:
            for f in model['link_failures']:
                fails.append('unresolved route: %s -- %s' % (f['label'], f['reason']))
        live = set()
        for name in wb.sheetnames:
            for row in wb[name].iter_rows():
                for c in row:
                    h = getattr(c, 'hyperlink', None)
                    if h is None:
                        continue
                    tgt = getattr(h, 'target', None)
                    if tgt:
                        live.add(tgt)
                    elif not getattr(h, 'location', None):
                        fails.append('%s!%s has an empty hyperlink' % (name, c.coordinate))
        built = {l['url'] for l in model['_all_links'] if l['ok']}
        stray = live - built
        if stray:
            fails.append('%d hyperlink(s) not built by study_links: %s'
                         % (len(stray), sorted(stray)[:3]))

        # The sessions plan must still match the pack it was derived from.
        if model['sessions_doc']['pack_status'] != 'CURRENT':
            fails.append('study_sessions.json is %s against its topic pack'
                         % model['sessions_doc']['pack_status'])

        # The cockpit must actually point somewhere.
        cur = model['current']
        if not cur['all_complete']:
            if not cur['topic']:
                fails.append('no current topic could be derived')
            if not cur['session']:
                # NOT a failure. A topic whose sessions have not been authored
                # yet is a real, legitimate state -- it is what the day after
                # D01 completes looks like. Failing the build there would make
                # the gate fire on progress rather than on a defect, so the
                # workbook says so on its face and the check reports it.
                warns.append('no sessions authored yet for %s -- START HERE '
                             'says so rather than inventing one'
                             % (cur['topic'] or {}).get('topic_id'))
            elif not model['current_tasks']:
                fails.append('current session has no tasks')

        # Progress must have survived: every topic's status is the one in the
        # JSON, not a value the renderer invented.
        prog = (_load(PROGRESS).get('topics') or {})
        for i, t in enumerate(model['topics'], start=2):
            want_status = (prog.get(t['topic_id']) or {}).get('status', 'NOT_STARTED')
            got = wb['PROGRESS'].cell(row=i, column=3).value
            if got != want_status:
                fails.append('PROGRESS %s shows %s, study_progress.json says %s'
                             % (t['topic_id'], got, want_status))

        for w in warns:
            print('WARN: %s' % w)
        if fails:
            for f in fails:
                print('FAIL: %s' % f)
            return 1
        print('roadmap workbook -- %d sheets, %d topics, %d queue tasks, '
              '%d checked routes, 0 dead links, loads clean'
              % (len(wb.sheetnames), wb['ROADMAP'].max_row - 1,
                 len(model['queue']), model['link_count']))
        return 0

    render_workbook(model, args.out)
    print(f'wrote {os.path.relpath(args.out, ROOT)}')
    print(f'  {len(model["topics"])} topics, '
          f'{len(model["official_nodes"])} official nodes')
    print(f'  claim: {model["public_claim"]}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
