#!/usr/bin/env python3
"""Controls proving the study system can grow without being redesigned.

The failure this suite exists to prevent is not a crash. It is the quieter
one: a schema that works perfectly today and has to be rewritten the moment
sixteen years of Written Question Intelligence arrive. So the tests assert
expandability directly --

  * a consumer must tolerate every historical-QI field being null;
  * a PARTIAL recovery must be storable as exactly the range recovered;
  * a fake "COMPLETE" claim must be REJECTED, not rounded up;
  * public copy must be derived, so it cannot outrun the stored evidence;
  * today's numbers must not move while all of the above is true.

Fixtures are synthetic and in-memory. Nothing here writes a governed artefact,
because a self-test that harvests live corpus state is a wasting asset -- it
passes until the corpus grows and then silently measures nothing.

Usage:  python tools/study/test_study_expandability.py
"""
import copy, io, json, os, re, sys

if __name__ == '__main__':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, HERE)

import evidence_model as EM
import export_roadmap_xlsx as RX
import build_topic_pages as BP

D = os.path.join(ROOT, 'docs', 'study')
PASS, FAIL = [], []


def ok(name, cond, detail=''):
    (PASS if cond else FAIL).append(name if cond else f'{name}: {detail}')


def _load(n):
    return json.load(open(os.path.join(D, n), encoding='utf-8'))


# --------------------------------------------------------------------------- #
# Backward compatibility -- today's numbers must not move
# --------------------------------------------------------------------------- #
def test_current_values_preserved():
    spine, mappings = _load('study_spine.json'), _load('study_mappings.json')
    official, cov = _load('official_syllabus.json'), _load('coverage_matrix.json')
    ok('10 canonical topics preserved', len(spine['domains']) == 10,
       str(len(spine['domains'])))
    ok('topic ids are still D01..D10',
       sorted(d['domain_id'] for d in spine['domains'])
       == [f'D{i:02d}' for i in range(1, 11)])
    ok('25 official nodes preserved', len(official['nodes']) == 25,
       str(len(official['nodes'])))
    ok('1081 mappings preserved', len(mappings['mappings']) == 1081,
       str(len(mappings['mappings'])))
    ok('every official node still scored for coverage',
       len(cov['nodes']) == 25, str(len(cov['nodes'])))
    ok('mapping summary still accounts for every record',
       sum(mappings['summary']['by_status'].values()) == 1081)


def test_stable_identity_fields_present():
    """Stable ids are load-bearing; evidence growth must not migrate them."""
    rec = next(iter(_load('study_mappings.json')['mappings'].values()))
    for f in ('canonical_question_id', 'content_type', 'topic_id',
              'syllabus_version', 'official_syllabus_version'):
        ok(f'stable identity field {f} present', f in rec)
    node = _load('official_syllabus.json')['nodes'][0]
    for f in ('official_node_id', 'official_order', 'source_digest'):
        ok(f'stable official field {f} present', f in node)


# --------------------------------------------------------------------------- #
# Expandability -- the historical QI socket
# --------------------------------------------------------------------------- #
def test_socket_is_declared_and_empty():
    h = _load('written_evidence_horizon.json')['layers']['historical_written_qi']
    ok('historical layer is declared', h['layer'] == 'HISTORICAL_WRITTEN_QI')
    ok('historical layer is honestly NOT_STARTED', h['status'] == 'NOT_STARTED',
       h['status'])
    ok('socket publishes its dormancy vocabulary',
       set(h['dormancy_classes']) == set(EM.DORMANCY_CLASSES))
    ok('socket publishes its relevance vocabulary',
       set(h['relevance_classes']) == set(EM.RELEVANCE))
    ok('socket distinguishes long-term from recent recurrence',
       'all_time' in h['recurrence_windows']
       and 'last_5_years' in h['recurrence_windows']
       and 'last_3_years' in h['recurrence_windows'],
       str(h['recurrence_windows']))
    ok('socket carries a stable family id namespace',
       h['family_id_namespace'] == 'FAMILY-EM-', h['family_id_namespace'])
    # The 2010-2020 window ceased to be "unreachable" on 2026-08-23 when the
    # source layer was adopted, so the single 2006-2020 gap correctly split in
    # two. What must never happen is either half quietly disappearing: a gap
    # that stops being recorded reads as coverage.
    gaps = h['known_gaps']
    ok('the 2006-2009 gap is recorded rather than glossed',
       any(g['from_year'] == 2006 and g['to_year'] == 2009 for g in gaps),
       str(gaps))
    ok('the 2010-2020 window is still recorded as a QI gap',
       any(g['from_year'] == 2010 and g['to_year'] == 2020 for g in gaps),
       str(gaps))
    ok('the 2010-2020 gap says the questions are NOT ingested',
       any(g['from_year'] == 2010 and 'NOT ingested' in g['reason']
           for g in gaps), str(gaps))
    ok('adopting the source layer did not populate the QI coverage layer',
       h['papers_total'] is None and h['questions_total'] is None,
       f"{h['papers_total']}/{h['questions_total']}")
    ok('the socket says only the SOURCE layer was adopted',
       h['source_status'] == 'SOURCE_LAYER_ADOPTED_QUESTIONS_NOT_INGESTED',
       h['source_status'])
    ok('an unpopulated coverage layer vouches for no dates',
       h['date_certainty'] == 'NONE', h['date_certainty'])


def test_consumers_tolerate_a_null_socket():
    """A renderer that breaks on an empty socket is not actually expandable."""
    model = RX.build_model()
    ok('roadmap model builds with the socket empty', len(model['topics']) == 10)
    for t in model['topics']:
        for f in RX.FUTURE_WRITTEN_FIELDS:
            ok(f'{t["topic_id"]} carries reserved field {f}', f in t)
    ok('reserved fields render as NOT YET INTEGRATED, never as a number',
       all(t['historical_written_papers'] == RX.NOT_YET
           for t in model['topics']))


def test_partial_recovery_is_storable_exactly():
    """If only 2018-2026 is recovered, store 2018-2026 -- not 2010-2026."""
    partial = EM.empty_qi_socket(
        'VALIDATED_RANGE', date_certainty='OFFICIAL_DATED',
        papers_total=88, questions_total=792,
        earliest_year=2018, latest_year=2026,
        earliest_sitting='2018-01', latest_sitting='2026-08',
        validated_ranges=[{'from_year': 2018, 'to_year': 2026}],
        known_gaps=[{'from_year': 2010, 'to_year': 2017,
                     'reason': 'not recovered'}])
    ok('a partial range validates', EM.assert_honest(partial) == [],
       str(EM.assert_honest(partial)))
    ok('a partial range keeps its real earliest year',
       partial['earliest_year'] == 2018)
    ok('a partial range still records what is missing',
       partial['known_gaps'][0]['from_year'] == 2010)


def test_fake_completeness_is_rejected():
    bad = EM.empty_qi_socket('VALIDATED_RANGE')          # no ranges, no counts
    ok('VALIDATED_RANGE with no evidence is rejected', EM.assert_honest(bad) != [])

    lying = EM.empty_qi_socket(
        'COMPLETE', papers_total=10, questions_total=90,
        earliest_year=2010, latest_year=2026,
        validated_ranges=[{'from_year': 2010, 'to_year': 2026}],
        known_gaps=[{'from_year': 2012, 'to_year': 2015, 'reason': 'missing'}])
    ok('COMPLETE while gaps are recorded is rejected',
       EM.assert_honest(lying) != [], 'accepted a contradictory completeness claim')

    inverted = EM.empty_qi_socket(
        'VALIDATED_RANGE', papers_total=5, questions_total=45,
        earliest_year=2026, latest_year=2010,
        validated_ranges=[{'from_year': 2010, 'to_year': 2026}])
    ok('an inverted year range is rejected', EM.assert_honest(inverted) != [])

    populated_but_unstarted = EM.empty_qi_socket('NOT_STARTED', papers_total=40)
    ok('NOT_STARTED carrying counts is rejected',
       EM.assert_honest(populated_but_unstarted) != [])

    try:
        EM.empty_qi_socket('MOSTLY_DONE')
        ok('an unknown status is rejected', False, 'accepted MOSTLY_DONE')
    except SystemExit:
        ok('an unknown status is rejected', True)


# --------------------------------------------------------------------------- #
# Public claim policy
# --------------------------------------------------------------------------- #
def test_public_claim_cannot_outrun_the_evidence():
    horizon = _load('written_evidence_horizon.json')
    cur = horizon['layers']['current_solved_written']
    hist = horizon['layers']['historical_written_qi']
    claim = EM.public_evidence_claim(cur, hist)
    ok('claim is the stored derived sentence',
       claim == horizon['public_claim']['derived_sentence'])
    for forbidden in ('2010', '16 years', 'since 2010'):
        ok(f'claim does not assert {forbidden!r}', forbidden not in claim, claim)
    ok('claim states the real current span',
       cur['earliest_sitting'] in claim and cur['latest_sitting'] in claim)

    # ...and it strengthens BY ITSELF once, and only once, evidence exists.
    validated = EM.empty_qi_socket(
        'VALIDATED_RANGE', date_certainty='OFFICIAL_DATED',
        papers_total=88, questions_total=792,
        earliest_year=2015, latest_year=2026,
        validated_ranges=[{'from_year': 2015, 'to_year': 2026}])
    stronger = EM.public_evidence_claim(cur, validated)
    ok('a validated layer strengthens the claim automatically',
       '2015' in stronger and 'historical' in stronger.lower(), stronger)
    ok('a NOT_STARTED layer never strengthens it',
       EM.public_evidence_claim(cur, hist) == claim)


def test_coverage_alone_never_licences_a_public_dated_claim():
    """The 2026-08-23 trap, pinned.

    The 2010-2020 recovery holds 115 of 132 months with every page hashed.
    Read on coverage alone that is nearly complete, and a future session could
    reasonably promote the socket and start saying "papers since 2010" in
    public. It would be wrong: not one of those sittings is dated by an
    official document. Coverage and date certainty are separate axes and the
    public claim hangs off the second one.
    """
    cur = _load('written_evidence_horizon.json')['layers']['current_solved_written']

    trap = EM.empty_qi_socket(
        'COMPLETE', date_certainty='SECONDARY_CLAIMED',
        papers_total=115, questions_total=1026,
        earliest_year=2010, latest_year=2020,
        validated_ranges=[{'from_year': 2010, 'to_year': 2020}])
    allowed, why = EM.date_certainty_gate(trap)
    ok('COMPLETE coverage on secondary dates is BARRED from a public claim',
       not allowed, why)
    ok('...and the derived sentence does not mention the historical span',
       '2010' not in EM.public_evidence_claim(cur, trap),
       EM.public_evidence_claim(cur, trap))

    for tier in ('NONE', 'SECONDARY_CLAIMED', 'MIXED'):
        s2 = EM.empty_qi_socket(
            'COMPLETE', date_certainty=tier, papers_total=115,
            questions_total=1026, earliest_year=2010, latest_year=2020,
            validated_ranges=[{'from_year': 2010, 'to_year': 2020}])
        ok(f'date_certainty {tier} is barred', not EM.date_certainty_gate(s2)[0])
    official = EM.empty_qi_socket(
        'COMPLETE', date_certainty='OFFICIAL_DATED', papers_total=115,
        questions_total=1026, earliest_year=2010, latest_year=2020,
        validated_ranges=[{'from_year': 2010, 'to_year': 2020}])
    ok('date_certainty OFFICIAL_DATED over a validated range is permitted',
       EM.date_certainty_gate(official)[0])
    ok('OFFICIAL_DATED still needs coverage',
       not EM.date_certainty_gate(EM.empty_qi_socket(
           'PARTIAL', date_certainty='OFFICIAL_DATED', papers_total=3,
           earliest_year=2010, latest_year=2011))[0])
    ok('the required tier is exactly OFFICIAL_DATED',
       EM.PUBLIC_DATED_CLAIM_REQUIRES == 'OFFICIAL_DATED')
    ok('an unknown date-certainty tier is rejected by the honesty check',
       EM.assert_honest(EM.empty_qi_socket(
           'PARTIAL', date_certainty='PROBABLY_FINE')) != [])
    ok('papers held with date_certainty NONE is rejected',
       EM.assert_honest(EM.empty_qi_socket(
           'PARTIAL', papers_total=115)) != [])


def test_adopted_source_layer_adds_no_coverage():
    """ADOPT_SOURCE_LAYER_ONLY must stay exactly that."""
    hsl = _load('historical_source_layer.json')
    ok('source layer is adopted', hsl['status'] == 'ADOPTED_SOURCE_LAYER_ONLY',
       hsl['status'])
    ok('source layer holds 115 archived pages',
       len(hsl['papers']) == 115, str(len(hsl['papers'])))
    ok('every page is re-obtainable',
       all(p.get('sha256') and p.get('archive_url') for p in hsl['papers']))
    ok('every sitting date is a secondary claim',
       all(p['date_certainty'] == 'MONTH_YEAR_CLAIMED_BY_SECONDARY_SOURCE'
           for p in hsl['papers']))
    ok('the three claims are kept apart',
       hsl['claim_separation']['A_question_text_existed_on_the_page']
       == 'CORROBORATED'
       and hsl['claim_separation']['B_question_belonged_to_that_sitting']
       == 'CLAIMED_BY_SECONDARY_SOURCE'
       and hsl['claim_separation']['C_officially_administered_in_that_sitting']
       == 'NOT_ESTABLISHED', str(hsl['claim_separation']))
    # The layer must carry provenance and NOT content.
    blob = json.dumps(hsl)
    for leaked in ('raw_wording', 'normalized_wording', 'text_verbatim',
                   'modern_corpus_joins', 'family_joins'):
        ok(f'no {leaked} leaked into the governed source layer',
           f'"{leaked}"' not in blob)
    ok('the research branch is recorded as unmerged',
       hsl['research_provenance']['merged_into_main'] is False)
    ok('the research commit is pinned, not floating',
       len(hsl['research_provenance']['commit']) == 40)
    # And it must not have moved a single MIW count.
    horizon = _load('written_evidence_horizon.json')
    cur = horizon['layers']['current_solved_written']
    ok('the current written corpus is untouched by the adoption',
       cur['papers_total'] == 40 and cur['questions_total'] == 360,
       f"{cur['papers_total']}/{cur['questions_total']}")
    ok('month gaps distinguish no-source-page from no-archive-capture',
       {g['classification'] for g in hsl['month_gaps']}
       == {'NO_SOURCE_PAGE', 'NO_ARCHIVE_CAPTURE'},
       str([g['classification'] for g in hsl['month_gaps']]))
    ok('no month gap is converted into a no-sitting claim',
       all(g['no_sitting_evidence'] != 'EVIDENCED' for g in hsl['month_gaps']))


def test_pages_make_no_unearned_claim():
    for name in ('topics.html', 'study.html'):
        page = open(os.path.join(ROOT, 'meoclass1', name),
                    encoding='utf-8').read()
        for forbidden in ('since 2010', '16 years', '2010-2026', '2010–2026'):
            ok(f'{name} makes no {forbidden!r} claim', forbidden not in page)
        ok(f'{name} states the adopted syllabus is not yet in force',
           '2027-01-01' in page, 'effective date missing')


# --------------------------------------------------------------------------- #
# Leakage
# --------------------------------------------------------------------------- #
def test_no_review_queue_leakage():
    """An unadjudicated mapping must not be presented as settled."""
    queue = {i['canonical_question_id'] for i in _load('mapping_review_queue.json')['items']}
    store = _load('study_mappings.json')['mappings']
    ok('every queued item is genuinely unsettled',
       all(store[q]['mapping_status'] != 'VALID_MAPPED'
           or store[q].get('last_reviewed') for q in queue if q in store),
       'a queued question is published as VALID_MAPPED without a review stamp')
    page = open(os.path.join(ROOT, 'meoclass1', 'topics.html'),
                encoding='utf-8').read()
    ok('the topic page never labels a mapping "confirmed" or "verified"',
       not re.search(r'\b(confirmed|verified) mapping\b', page, re.I))


def test_no_paid_answer_leakage():
    """The study pages carry question stems and links -- never answers."""
    page = open(os.path.join(ROOT, 'meoclass1', 'topics.html'),
                encoding='utf-8').read()
    for marker in ('15-second answer', '60-second answer', 'model_answer',
                   'CE Oral Tip', 'class="answer"', 'Exam Plan'):
        ok(f'topic page carries no {marker!r}', marker not in page)
    study = open(os.path.join(ROOT, 'meoclass1', 'study.html'),
                 encoding='utf-8').read()
    # The shared stylesheet defines .q-list, so the marker to look for is a
    # rendered list, not the class name.
    ok('study landing renders no question list at all',
       '<ul class="q-list">' not in study)
    ok('study landing links to no question anchor',
       not re.search(r'href="QB[^"]*#q\d+"', study))
    for name, text in (('topics.html', page), ('study.html', study)):
        ok(f'{name} does not link into the other paid product',
           '/solvedQP/' not in text,
           'cross-product link would bounce a customer to login')


# --------------------------------------------------------------------------- #
# Renderers
# --------------------------------------------------------------------------- #
def test_workbook_and_pages_build():
    wb_path = os.path.join(D, 'MIW_MEO_Class1_Study_Roadmap.xlsx')
    ok('roadmap workbook exists', os.path.exists(wb_path))
    if os.path.exists(wb_path):
        from openpyxl import load_workbook
        wb = load_workbook(wb_path)
        ok('workbook leads with the operational cockpit, then the projections',
           wb.sheetnames == ['START HERE', 'STUDY QUEUE', 'ROADMAP',
                             'TOPIC DETAIL', 'WRITTEN BY TOPIC',
                             'OFFICIAL SYLLABUS', 'COVERAGE', 'WRITTEN QI',
                             'PROGRESS', 'ABOUT'],
           str(wb.sheetnames))
        ok('START HERE is the first sheet a candidate sees',
           wb.sheetnames[0] == 'START HERE', str(wb.sheetnames[:1]))
        ok('ROADMAP lists every topic', wb['ROADMAP'].max_row == 11,
           str(wb['ROADMAP'].max_row))
        ok('OFFICIAL SYLLABUS lists every official item',
           wb['OFFICIAL SYLLABUS'].max_row == 26,
           str(wb['OFFICIAL SYLLABUS'].max_row))
        headers = [c.value for c in wb['ROADMAP'][1]]
        ok('roadmap separates study order from priority rank',
           'Study Order' in headers and 'Priority Rank' in headers, str(headers))


def test_study_order_is_dependency_first():
    """Score order would open the candidate on a topic gated behind D01."""
    spine = _load('study_spine.json')
    order = RX.study_order(spine['domains'])
    by_id = {d['domain_id']: d for d in spine['domains']}
    ok('D01 is first in study order', order['D01'] == 1, str(order))
    for did, d in by_id.items():
        for p in d['prerequisites']:
            ok(f'{did} comes after its prerequisite {p}', order[p] < order[did])
    top = min(by_id.values(), key=lambda d: d['priority_rank'])['domain_id']
    ok('the top-ranked topic is NOT assumed to be first to study',
       top != 'D01' and order[top] > 1,
       'this control is vacuous if the two orders ever coincide')


def test_progress_survives_regeneration():
    progress = _load('study_progress.json')
    ok('progress is hand-maintained, not generated',
       'HAND-MAINTAINED' in progress['authority'])
    ok('progress declares the full progression',
       'NOT_STARTED' in progress['progression']
       and 'MOCK_READY' in progress['progression'])
    # A topic absent from the file must default, not explode.
    stripped = copy.deepcopy(progress)
    stripped['topics'] = {}
    saved = RX._load
    try:
        RX._load = lambda p: stripped if p.endswith('study_progress.json') else saved(p)
        model = RX.build_model()
        ok('a topic missing from progress defaults to NOT_STARTED',
           all(t['study_status'] == 'NOT_STARTED' for t in model['topics']))
    finally:
        RX._load = saved
    # Progress must be an INPUT only. Assert no writer ever opens it.
    for gen in ('export_roadmap_xlsx.py', 'build_topic_pages.py',
                'build_study_spine.py', 'build_study_mappings.py'):
        src = open(os.path.join(HERE, gen), encoding='utf-8').read()
        ok(f'{gen} never opens the progress file for writing',
           not re.search(r"open\(\s*PROGRESS\s*,\s*['\"]w", src)
           and "study_progress.json'), 'w'" not in src)


def test_public_roadmap_is_derived_and_bounded():
    """The PUBLIC projection: right shape, right numbers, nothing paid.

    Structural checks alone would pass a generator whose guards had been
    deleted, so every guard is also MUTATED here: the control is not "the page
    is clean today" but "the thing that keeps it clean actually fires".
    """
    import build_public_study_roadmap as PR
    spine = _load('study_spine.json')
    p = PR.project()
    doc = PR.render(p)
    PR.assert_public_safe(doc, p)

    # --- shape ----------------------------------------------------------
    ids = [t['topic_id'] for t in p['topics']]
    ok('public page shows 10 topics', len(ids) == 10, str(len(ids)))
    ok('public topic ids are unique and are D01..D10',
       sorted(ids) == [f'D{i:02d}' for i in range(1, 11)], str(sorted(ids)))

    order = RX.study_order(spine['domains'])
    ok('public study order is the governed study order',
       all(t['study_order'] == order[t['topic_id']] for t in p['topics']),
       str({t['topic_id']: t['study_order'] for t in p['topics']}))
    ok('D01 leads the public roadmap', p['topics'][0]['topic_id'] == 'D01')

    # --- counts are read, never typed -----------------------------------
    by_id = {d['domain_id']: d for d in spine['domains']}
    for t in p['topics']:
        d = by_id[t['topic_id']]
        ok(f'{t["topic_id"]} oral count matches the spine',
           t['oral_questions'] == d['oral']['questions'])
        ok(f'{t["topic_id"]} written count matches the spine',
           t['current_written_questions'] == d['written']['questions'])
        ok(f'{t["topic_id"]} examiner evidence matches the spine',
           t['examiner_evidenced_oral']
           == d['examiner_intelligence']['oral_questions_with_evidence'])
        for num in (t['oral_questions'], t['current_written_questions'],
                    t['official_syllabus_items']):
            ok(f'{t["topic_id"]} renders its derived figure {num}',
               f'>{num}</div>' in doc)

    # --- the evidence horizon is stated, not exceeded --------------------
    cur = _load('written_evidence_horizon.json')['layers']['current_solved_written']
    ok('the public evidence sentence is the derived one',
       p['evidence_sentence'] == EM.public_evidence_claim(
           cur, _load('written_evidence_horizon.json')
           ['layers']['historical_written_qi']))
    ok('the current sitting range is stated on the page',
       cur['earliest_sitting'] in doc and cur['latest_sitting'] in doc)
    ok('historical QI is not claimed as integrated',
       not p['historical_integrated'] and 'being expanded' in doc)

    # --- routing --------------------------------------------------------
    hrefs = set(re.findall(r'href="([^"]+)"', doc))
    for route in (PR.ROUTE_ORAL_TRIAL, PR.ROUTE_WRITTEN_TRIAL, PR.ROUTE_STORE,
                  PR.ROUTE_EXAMINER, PR.ROUTE_ORAL_SAMPLE, PR.ROUTE_WRITTEN_SAMPLE):
        ok(f'public route present: {route}', route in hrefs)
        target = route.split('?')[0].lstrip('/')
        ok(f'route target exists on disk: {target}',
           os.path.exists(os.path.join(ROOT, target)))
    ok('no href reaches a gated product surface',
       not any(h.startswith(('/meoclass1', '/solvedQP')) for h in hrefs),
       str(sorted(hrefs)))

    # --- discovery surface ----------------------------------------------
    ok('GA4 is the canonical property, installed once',
       doc.count(PR.GA4) == 2)
    for name in ('study_roadmap_view', 'study_topic_preview', 'study_oral_cta',
                 'study_written_cta'):
        ok(f'GA4 event wired: {name}', name in doc)
    ok('the public page is indexable', 'content="index, follow"' in doc)
    ok('SEO metadata is present',
       '<title>' in doc and 'name="description"' in doc
       and 'rel="canonical"' in doc and 'og:title' in doc)

    # --- MUTATION: every guard must actually fire -------------------------
    def fires(label, mutate_doc=None, mutate_p=None):
        d2 = mutate_doc(doc) if mutate_doc else doc
        p2 = mutate_p(copy.deepcopy(p)) if mutate_p else p
        try:
            PR.assert_public_safe(d2, p2)
        except PR.Unsafe:
            ok(f'guard fires: {label}', True)
        else:
            ok(f'guard fires: {label}', False, 'mutation survived')

    fires('a link into the gated Oral product',
          lambda d: d + '<a href="/meoclass1/QB1_A.html">x</a>')
    fires('a link into the gated Written product',
          lambda d: d + '<a href="/solvedQP/index.html">x</a>')
    fires('an unsupported 16-year corpus claim',
          lambda d: d + '<p>Based on 16 years of papers.</p>')
    fires('a 2010-2026 span claim', lambda d: d + '<p>2010-2026 analysis</p>')
    fires('an answer-surface marker', lambda d: d + '<p>CE Oral Tip: open it.</p>')
    fires('a cheatsheet reference', lambda d: d + '<a>CheatSheet</a>')
    fires('an internal review token', lambda d: d + '<p>REVIEW_PENDING</p>')
    fires("a leaked private field", mutate_p=lambda q: (
        q['topics'][0].__setitem__('study_status', 'IN_PROGRESS') or q))
    fires('a leaked priority score', mutate_p=lambda q: (
        q['topics'][3].__setitem__('priority_score', 0.9) or q))
    fires('a sample quota overrun', mutate_p=lambda q: (
        q['topics'][0].__setitem__(
            'samples', q['topics'][0]['samples'] + ['a', 'b']) or q))
    fires('a missing topic', mutate_p=lambda q: (q['topics'].pop(), q)[1])
    fires('a duplicated topic', mutate_p=lambda q: (
        q['topics'].__setitem__(1, q['topics'][0]) or q))

    # --- MUTATION: the page widens itself when the socket is filled -------
    horizon = _load('written_evidence_horizon.json')
    widened = copy.deepcopy(horizon)
    widened['layers']['historical_written_qi'].update(
        status='VALIDATED_RANGE', date_certainty='OFFICIAL_DATED',
        earliest_year=2011, latest_year=2020,
        papers_total=70, questions_total=630)
    # build_evidence_horizon.py rewrites the stored sentence whenever the
    # socket changes, so the fixture does the same. Skipping this would test a
    # state the real toolchain never produces -- and would instead exercise the
    # drift guard, which has its own control below.
    widened['public_claim']['derived_sentence'] = EM.public_evidence_claim(
        widened['layers']['current_solved_written'],
        widened['layers']['historical_written_qi'])
    # BOTH loaders must be patched. The public sentence is not written by the
    # public generator -- it is inherited from RX.build_model(), which reads
    # the horizon through its OWN loader. Patching only PR._load would widen
    # the socket the page reports while leaving the sentence it prints on the
    # old evidence, which is precisely the drift this control exists to catch.
    saved, saved_rx = PR._load, RX._load
    try:
        PR._load = (lambda n: widened
                    if n == 'written_evidence_horizon.json' else saved(n))
        RX._load = (lambda path: widened
                    if path.endswith('written_evidence_horizon.json')
                    else saved_rx(path))
        p2 = PR.project()
        doc2 = PR.render(p2)
        ok('a validated historical range strengthens the public sentence',
           '2011' in p2['evidence_sentence'] and '2020' in p2['evidence_sentence'],
           p2['evidence_sentence'])
        ok('the page stops saying the layer is being expanded',
           'being expanded' not in doc2 and 'is integrated' in doc2)
        ok('no marketing copy had to be rewritten for that transition',
           doc2.count('MEO Class I <span>Study Roadmap</span>') == 1)
        # And the drift guard itself: a socket widened WITHOUT regenerating the
        # stored sentence must stop the build rather than publish either wording.
        stale = copy.deepcopy(widened)
        stale['public_claim']['derived_sentence'] = horizon['public_claim'][
            'derived_sentence']
        PR._load = (lambda n: stale
                    if n == 'written_evidence_horizon.json' else saved(n))
        RX._load = (lambda path: stale
                    if path.endswith('written_evidence_horizon.json')
                    else saved_rx(path))
        try:
            PR.project()
            ok('a stale stored evidence sentence stops the build', False,
               'drift was published')
        except PR.Unsafe:
            ok('a stale stored evidence sentence stops the build', True)
    finally:
        PR._load, RX._load = saved, saved_rx

    # Samples must be real corpus text, never invented for the page.
    mappings = _load('study_mappings.json')['mappings']
    corpus = {(r.get('text') or '').strip() for r in mappings.values()}
    for t in p['topics']:
        for s in t['samples']:
            ok(f'{t["topic_id"]} sample is verbatim corpus text', s in corpus,
               s[:60])
    total = sum(len(t['samples']) for t in p['topics'])
    ok('sampled stems stay a small fraction of the oral corpus',
       total <= PR.SAMPLES_PER_TOPIC * 10
       and total < spine['totals']['oral_questions_total'] * 0.1, str(total))


def main():
    for fn in (test_current_values_preserved,
               test_stable_identity_fields_present,
               test_socket_is_declared_and_empty,
               test_consumers_tolerate_a_null_socket,
               test_partial_recovery_is_storable_exactly,
               test_fake_completeness_is_rejected,
               test_public_claim_cannot_outrun_the_evidence,
               test_coverage_alone_never_licences_a_public_dated_claim,
               test_adopted_source_layer_adds_no_coverage,
               test_pages_make_no_unearned_claim,
               test_no_review_queue_leakage,
               test_no_paid_answer_leakage,
               test_workbook_and_pages_build,
               test_study_order_is_dependency_first,
               test_public_roadmap_is_derived_and_bounded,
               test_progress_survives_regeneration):
        fn()
    print(f'study expandability -- {len(PASS) + len(FAIL)} assertions')
    for f in FAIL:
        print('  FAIL ' + f)
    if FAIL:
        print(f'\n{len(FAIL)} FAILED')
        return 1
    print(f'  all {len(PASS)} PASS')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
