"""Export the CURRENT published MIW Oral question bank to a candidate workbook.

Born 2026-08-22 for the August 2026 INTERIM snapshot, so that every future
interim/final spreadsheet comes from repository truth rather than manual Excel
patching of the July workbook.

Architecture (and the one rule that keeps it honest):

    CANONICAL ORAL DATA        meoclass1/qb_content_index.json   (identity + text)
                               EXAMINER_INDEX_SNAPSHOT.json      (examiner rows)
                               live QB HTML                      (anchor proof only)
            |
    NORMALIZED EXPORT MODEL    build_export_model()  -> list of plain dicts
            |
    XLSX RENDERER              render_workbook()     -> openpyxl workbook

The renderer NEVER infers topic, examiner or syllabus data. Every cell is a
field of the model, and every field of the model is read from a governed
repository artefact. The model already carries the syllabus fields the future
governed mapper will populate (official_syllabus_version, official_syllabus_node_id,
miw_topic_id, miw_topic_name, objective_id); today they are empty and are NOT
rendered -- this exporter must never populate them itself.

Sources and what each is trusted for:
  * qb_content_index.json  - the 738/86 canonical identity (file + anchor), the
                             candidate-facing question text, the QB group and the
                             page title used as the current production
                             "Topic / Category". The exporter re-runs the index
                             generator's --check so a stale index cannot be exported.
  * EXAMINER_INDEX_SNAPSHOT.json - exactly the object meoclass1/examiner-index.html
                             is rendered from. One row = one (examiner, question)
                             relationship (958 / 7 examiners on production). The
                             exporter joins rows to questions by canonical id and
                             shows examiner NAMES only; tiers, refs and evidence
                             counts stay internal.
  * live QB HTML           - proof that each exported anchor exists on the page it
                             links to (id="q<N>" on a q-card in that file).

The group-facing workbook additionally carries a "<Month> <Year> - New & Updated"
sheet, placed second so it is the first thing a candidate sees. Its rows are a
projection of tools/oral/oral_monthly.py over governed provenance -- the corpus
as it stood at the last commit before the month opened, compared with the corpus
at the last commit before it closed, cross-checked against the batch and
correction manifests DATED INSIDE THAT MONTH. The month is bounded at BOTH ends:
a workbook exported in September must not sell September's work as August's.
Nothing on that sheet is hand-curated, and a row that cannot name its evidence
fails the export.

Usage:
  PYTHONIOENCODING=utf-8 python tools/oral/export_question_bank_xlsx.py --candidate-share
      writes docs/MIW-master-Question-bank/MIW_MEO_Class1_Oral_QuestionBank_<Month>_<Year>.xlsx
  ... --month YYYY-MM       which month the New & Updated sheet projects
  ... --candidate-interim   also rewrite MIW_August2026_QuestionBank_INTERIM.xlsx
  ... --out PATH            write somewhere else (a validator/test scratch path)
  ... --working-master      also write MEO_QB_master_v27_WORKING.xlsx (internal,
                            same model, adds the Examiner-relationship projection)
  ... --model-json PATH     dump the normalized model (for diffing / tests)

Historical release workbooks are never overwritten: any output whose file name
is in PROTECTED_OUTPUTS (or ends in a FINAL-style name) is refused.
"""
from __future__ import annotations

import argparse
import datetime as _dt
import html as _html
import json
import os
import re
import subprocess
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import oral_lib as L  # noqa: E402
import oral_monthly as MM  # noqa: E402

SITE_BASE = "https://marineintelligenceweekly.com"
INDEX_PATH = L.MEO / "qb_content_index.json"
SNAPSHOT_PATH = L.OUT / "EXAMINER_INDEX_SNAPSHOT.json"
REGISTER_PATH = L.OUT / "EXAMINER_ALIAS_REGISTER.json"
RELEASE_DIR = L.REPO / "docs" / "MIW-master-Question-bank"

INTERIM_NAME = "MIW_August2026_QuestionBank_INTERIM.xlsx"
WORKING_NAME = "MEO_QB_master_v27_WORKING.xlsx"

# The group-facing product. "INTERIM" is honest internally but reads as
# provisional to a candidate, and the workbook is a CURRENT SNAPSHOT rather than
# a frozen edition, so the share name carries the month and the About sheet
# carries the date -- neither declares the month's governance closed.
SHARE_NAME_FMT = "MIW_MEO_Class1_Oral_QuestionBank_%s_%s.xlsx"   # Month, YYYY
DEFAULT_MONTH = "2026-08"

# Historical releases (never overwritten) and the final names this exporter is
# not allowed to produce until the final release sequence authorises them.
PROTECTED_OUTPUTS = {
    "MEO_QB_master_v26.xlsx",
    "MIW_July2026_QuestionBank_SHARE.xlsx",
    "MEO_QB_master_v27.xlsx",
    "MIW_August2026_QuestionBank_SHARE.xlsx",
}
FINAL_NAME_RE = re.compile(r"final|master_v\d+\.xlsx$|_SHARE\.xlsx$", re.I)

SNAPSHOT_LABEL = "August 2026 — INTERIM SNAPSHOT"
ABOUT_NOTE = (
    "This is an interim snapshot of the current live MIW Oral question bank. "
    "Recent examiner follow-ups are still being incorporated and reconciled. "
    "A final audited consolidated workbook will follow. The current file is being "
    "shared now so candidates do not have to wait."
)
ACCESS_NOTE = "Question links use normal MIW access permissions."

# --- group-facing wording -------------------------------------------------
# Every claim here has to be one the repository can prove. In particular the
# month sheet describes what MIW ADDED AND UPDATED, never "everything asked in
# August": the corpus is built from the candidate reports MIW actually receives,
# and no artefact in this repository can support the stronger claim.
SHARE_LABEL_FMT = "%s %s — current snapshot"
SHARE_NOTE = (
    "The MIW MEO Class I Oral Question Bank is maintained continuously from "
    "candidate reports, governed editorial review and current-source verification. "
    "This workbook is the question index: every question below is live on MIW, and "
    "each row links straight to its page."
)
SHARE_ACCESS_NOTE = (
    "Question links open the MIW page for that question. Reading the full worked "
    "answer needs MIW access; the question index itself is yours to keep and study "
    "from either way."
)
MONTH_SHEET_FMT = "%s %s - New & Updated"
MONTH_HEADLINE_FMT = "MIW MEO Class I Oral Question Bank — %s %s: new and updated"
MONTH_BLURB = ("New and updated questions added to MIW during %s %s from governed "
               "candidate reports and editorial review.")
MONTH_LEGEND = ("NEW = not in the bank when the month opened.   "
                "UPDATED = already in the bank, materially revised this month.")
MONTH_LABELS = {MM.STATUS_NEW: "Added %s %s", MM.STATUS_UPDATED: "Updated %s %s"}

# Candidate sheet columns: (header, model field, width)
MAIN_COLUMNS = [
    ("No.", "no", 6),
    ("MIW Question ID", "id", 16),
    ("Question", "question", 80),
    ("Topic / Category", "topic", 34),
    ("Question Bank", "qb", 14),
    ("Examiner(s)", "examiners_text", 26),
    ("MIW Answer Link", "url", 52),
]
MAIN_SHEET = "All Questions"
ABOUT_SHEET = "About"

# Month sheet columns. Same fields as the main sheet plus the status, so a
# candidate reads one grammar across the workbook.
MONTH_COLUMNS = [
    ("Status", "month_status", 10),
    ("This Month", "month_label", 16),
    ("MIW Question ID", "id", 16),
    ("Question", "question", 80),
    ("Topic / Category", "topic", 34),
    ("Question Bank", "qb", 14),
    ("Examiner(s)", "examiners_text", 26),
    ("MIW Answer Link", "url", 52),
]
# The row of the month sheet that carries the table headers; the summary block
# above it is plain label/value cells, never a merged-cell banner.
MONTH_HEADER_ROW = 7

# Reserved for the governed syllabus mapper. Present in the model, never
# rendered to a candidate, never populated here.
SYLLABUS_FIELDS = (
    "official_syllabus_version",
    "official_syllabus_node_id",
    "miw_topic_id",
    "miw_topic_name",
    "objective_id",
)


class ExportFailure(Exception):
    pass


def fail(msg):
    raise ExportFailure(msg)


# ------------------------------------------------------------------ text

_WS_RE = re.compile(r"\s+")


def display_text(s):
    """Presentation-only normalisation: entities, markup, line breaks, whitespace.
    No semantic rewriting."""
    s = _html.unescape(L.strip_tags(s or ""))
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    return _WS_RE.sub(" ", s).strip()


def natural_file_key(fname):
    parts = re.split(r"(\d+)", fname)
    return [int(p) if p.isdigit() else p for p in parts]


# ------------------------------------------------------------------ sources

def check_index_fresh():
    """Refuse to export from an index that no longer matches the live HTML."""
    env = dict(os.environ, PYTHONIOENCODING="utf-8")
    r = subprocess.run([sys.executable, str(HERE / "build_qb_content_index.py"), "--check"],
                       capture_output=True, text=True, encoding="utf-8", env=env)
    if r.returncode != 0:
        fail("qb_content_index.json is stale against the live QB HTML:\n" + r.stdout + r.stderr)


def load_index():
    d = json.loads(INDEX_PATH.read_text(encoding="utf-8"))
    if d.get("manifest_version") != "1.1":
        fail("unexpected qb_content_index manifest_version %r" % d.get("manifest_version"))
    return d


def load_examiner_rows():
    s = json.loads(SNAPSHOT_PATH.read_text(encoding="utf-8"))
    reg = json.loads(REGISTER_PATH.read_text(encoding="utf-8"))
    names = {e["slug"]: e["canonical_name"] for e in reg["examiners"]}
    by_q = {}
    for r in s["rows"]:
        if r["slug"] not in names:
            fail("examiner slug %r in snapshot but not in the alias register" % r["slug"])
        by_q.setdefault(r["canonical_question_id"], []).append(r)
    return s, by_q, names


_ANCHOR_CACHE = {}


def anchor_exists(fname, anchor):
    """True when the live page carries a q-card with id=<anchor>."""
    if fname not in _ANCHOR_CACHE:
        p = L.MEO / fname
        if not p.is_file():
            _ANCHOR_CACHE[fname] = None
        else:
            h = p.read_text(encoding="utf-8")
            _ANCHOR_CACHE[fname] = set(re.findall(
                r'<div[^>]*\bclass="[^"]*\bq-card\b[^"]*"[^>]*\bid="(q\d+)"', h))
            _ANCHOR_CACHE[fname] |= set(re.findall(
                r'<div[^>]*\bid="(q\d+)"[^>]*\bclass="[^"]*\bq-card\b', h))
    ids = _ANCHOR_CACHE[fname]
    return ids is not None and anchor in ids


# ------------------------------------------------------------------ model

def month_names(month):
    """('2026-08') -> ('August', '2026')."""
    start, _ = MM.month_bounds(month)
    return start.strftime("%B"), start.strftime("%Y")


def build_export_model(check_fresh=True, month=None):
    """Return (meta, rows). Rows are plain dicts; one row per canonical question,
    in QB file order then document order. Raises ExportFailure on any identity
    defect so a broken model never reaches a renderer.

    With `month` (YYYY-MM) every row also carries month_status / month_label /
    month_evidence from oral_monthly. The evidence is the internal proof that a
    row belongs on the month sheet; it is a model field and is never rendered."""
    if check_fresh:
        check_index_fresh()
    idx = load_index()
    snapshot, ex_by_q, ex_names = load_examiner_rows()
    order = sorted(snapshot.get("sections", []), key=lambda s: s.get("slug", ""))
    slug_rank = {s["slug"]: i for i, s in enumerate(
        sorted(order, key=lambda s: -s.get("count", 0)))}

    rows, seen = [], set()
    files = sorted(idx["files"].items(), key=lambda kv: natural_file_key(kv[0]))
    for fname, f in files:
        if not (L.MEO / fname).is_file():
            fail("%s listed in the index but missing on disk" % fname)
        for q in f["questions"]:
            qid = q["id"]
            if qid in seen:
                fail("duplicate canonical id %s" % qid)
            seen.add(qid)
            if qid != "%s#%s" % (Path(fname).stem, q["anchor"]):
                fail("%s: id does not match file+anchor" % qid)
            if not anchor_exists(fname, q["anchor"]):
                fail("%s: anchor %s not found on %s" % (qid, q["anchor"], fname))
            text = display_text(q["text"])
            if not text:
                fail("%s: empty question text" % qid)
            exs = sorted({r["slug"] for r in ex_by_q.get(qid, [])},
                         key=lambda s: (slug_rank.get(s, 99), s))
            rows.append({
                "no": len(rows) + 1,
                "id": qid,
                "file": fname,
                "anchor": q["anchor"],
                "qnum": q["qnum"],
                "question": text,
                "qb": f["qb_group"],
                "topic": display_text(f["title"]),
                "url": "%s/meoclass1/%s#%s" % (SITE_BASE, fname, q["anchor"]),
                "examiners": [ex_names[s] for s in exs],
                "examiners_text": ", ".join(ex_names[s] for s in exs),
                "examiner_relationships": len(ex_by_q.get(qid, [])),
                "month_status": "",
                "month_label": "",
                "month_evidence": [],
                **{k: "" for k in SYLLABUS_FIELDS},
            })

    if len(rows) != idx["total_questions"]:
        fail("row count %d != index total_questions %d" % (len(rows), idx["total_questions"]))
    phantom = set(ex_by_q) - seen
    if phantom:
        fail("examiner snapshot names %d question ids not in the canonical index: %s"
             % (len(phantom), sorted(phantom)[:5]))

    if month:
        mname, myear = month_names(month)
        try:
            proj = MM.classify(month, {r["id"]: r["question"] for r in rows})
        except MM.MonthlyFailure as e:
            fail("monthly provenance for %s: %s" % (month, e))
        for r in rows:
            st = proj["status"].get(r["id"], "")
            r["month_status"] = st
            r["month_label"] = (MONTH_LABELS[st] % (mname, myear)) if st else ""
            r["month_evidence"] = proj["evidence"].get(r["id"], [])
        ungoverned = [r["id"] for r in rows if r["month_status"] and not r["month_evidence"]]
        if ungoverned:
            fail("%d month rows carry no provenance evidence: %s"
                 % (len(ungoverned), ungoverned[:5]))

    meta = {
        "snapshot_label": SNAPSHOT_LABEL,
        "canonical_questions": len(rows),
        "question_files": len({r["file"] for r in rows}),
        "examiner_relationships": snapshot["totals"]["relationships"],
        "examiners": snapshot["totals"]["examiners"],
        "questions_with_examiner": sum(1 for r in rows if r["examiners"]),
        "examiner_names": [ex_names[s["slug"]] for s in snapshot["sections"]],
        "index_source": str(INDEX_PATH.relative_to(L.REPO)).replace("\\", "/"),
    }
    if month:
        mname, myear = month_names(month)
        meta["month"] = month
        meta["month_name"], meta["month_year"] = mname, myear
        meta["month_sheet"] = MONTH_SHEET_FMT % (mname, myear)
        meta["month_new"] = sum(1 for r in rows if r["month_status"] == MM.STATUS_NEW)
        meta["month_updated"] = sum(1 for r in rows if r["month_status"] == MM.STATUS_UPDATED)
        meta["month_rows"] = meta["month_new"] + meta["month_updated"]
        meta["month_baseline_commit"] = proj["baseline_commit"]
        meta["month_baseline_questions"] = proj["baseline_questions"]
        if len(meta["month_sheet"]) > 31:
            fail("month sheet name %r exceeds Excel's 31-character limit"
                 % meta["month_sheet"])
    return meta, rows


# ------------------------------------------------------------------ renderer

def _styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    thin = Side(style="thin", color="D9D9D9")
    return {
        "head_font": Font(bold=True, color="FFFFFF"),
        "head_fill": PatternFill("solid", fgColor="1F3A5F"),
        "title_font": Font(bold=True, size=14, color="1F3A5F"),
        "bold": Font(bold=True),
        "link": Font(color="0563C1", underline="single"),
        "wrap": Alignment(wrap_text=True, vertical="top"),
        "top": Alignment(vertical="top"),
        "border": Border(top=thin, bottom=thin, left=thin, right=thin),
    }


def _table(ws, headers, widths, rows, st, link_col=None, header_row=1):
    """Render one header row + data rows, frozen and filtered. header_row lets a
    sheet carry a plain label/value summary above the table (the month sheet)
    without a merged-cell banner, which mobile Excel handles badly."""
    from openpyxl.utils import get_column_letter
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font, cell.fill, cell.alignment = st["head_font"], st["head_fill"], st["top"]
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    for i, r in enumerate(rows):
        for c, v in enumerate(r, 1):
            ws.cell(row=header_row + 1 + i, column=c, value=v)
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = st["wrap"]
            cell.border = st["border"]
        if link_col is not None:
            cell = row[link_col]
            cell.hyperlink = cell.value
            cell.font = st["link"]
    ws.freeze_panes = "A%d" % (header_row + 1)
    ws.auto_filter.ref = "A%d:%s%d" % (header_row, get_column_letter(len(headers)), ws.max_row)


def render_workbook(meta, rows, generated_at, internal=False):
    import openpyxl
    st = _styles()
    wb = openpyxl.Workbook()
    month = meta.get("month")

    # --- About (first sheet)
    ws = wb.active
    ws.title = ABOUT_SHEET
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90
    if month:
        label = SHARE_LABEL_FMT % (meta["month_name"], meta["month_year"])
        note, access = SHARE_NOTE, SHARE_ACCESS_NOTE
    else:
        label, note, access = meta["snapshot_label"], ABOUT_NOTE, ACCESS_NOTE
    lines = [
        ("Marine Intelligence Weekly", None, st["title_font"]),
        ("MEO Class I Oral Question Bank", None, st["bold"]),
        (label, None, st["bold"]),
        ("", None, None),
        ("Generated:", generated_at, None),
        ("Canonical questions:", meta["canonical_questions"], None),
        ("Question bank pages:", meta["question_files"], None),
        ("Examiners:", ", ".join(meta["examiner_names"]), None),
    ]
    if month:
        lines += [
            ("New this month:", meta["month_new"], None),
            ("Updated this month:", meta["month_updated"], None),
        ]
    lines += [
        ("", None, None),
        ("Note:", note, None),
        ("Access:", access, None),
        ("", None, None),
    ]
    if month:
        lines += [
            ("Start here:", "'%s' is what changed this month. 'All Questions' is the "
                            "complete current bank." % meta["month_sheet"], None),
            ("How to use:", "Filter any sheet by Topic / Category, Question Bank or "
                            "Examiner(s), then click the MIW Answer Link to open that "
                            "question on MIW.", None),
            ("Keeping current:", "The bank grows every month as candidates report what "
                                 "they were actually asked. Re-download the workbook for "
                                 "the current month's additions.", None),
        ]
    else:
        lines += [
            ("How to use:", "Open the 'All Questions' sheet. Filter by Topic / Category, "
                            "Question Bank or Examiner(s), then click the MIW Answer Link.", None),
        ]
    if internal:
        lines.append(("Internal:", "WORKING copy - not a release. Examiner relationships: %d "
                      "across %d examiners; %d of %d questions carry at least one examiner."
                      % (meta["examiner_relationships"], meta["examiners"],
                         meta["questions_with_examiner"], meta["canonical_questions"]), None))
    for a, b, font in lines:
        ws.append([a, b])
        if font:
            ws.cell(row=ws.max_row, column=1).font = font
        ws.cell(row=ws.max_row, column=2).alignment = st["wrap"]

    # --- <Month> - New & Updated: second sheet, so a candidate opening the file
    # sees this month's additions before the 738-row full bank.
    if month:
        ws = wb.create_sheet(meta["month_sheet"])
        # Plain single-cell lines in column A: no merges, and with the
        # neighbouring cells empty Excel lets each line overflow across the
        # sheet, which reads correctly on desktop, mobile and Sheets alike.
        summary = [
            (MONTH_HEADLINE_FMT % (meta["month_name"], meta["month_year"]), st["title_font"]),
            (MONTH_BLURB % (meta["month_name"], meta["month_year"]), None),
            ("Questions in the bank: %d        New this month: %d        "
             "Updated this month: %d" % (meta["canonical_questions"],
                                         meta["month_new"], meta["month_updated"]), st["bold"]),
            (MONTH_LEGEND, None),
            ("Current through: %s" % generated_at, None),
        ]
        for text, font in summary:
            ws.append([text])
            cell = ws.cell(row=ws.max_row, column=1)
            cell.alignment = st["top"]
            if font:
                cell.font = font
        if ws.max_row + 2 != MONTH_HEADER_ROW:
            fail("month summary block is %d rows; MONTH_HEADER_ROW says %d"
                 % (ws.max_row, MONTH_HEADER_ROW))
        m_headers = [h for h, _, _ in MONTH_COLUMNS]
        m_fields = [f for _, f, _ in MONTH_COLUMNS]
        # NEW before UPDATED, corpus order inside each; one row per card, so a
        # card touched twice this month still appears exactly once.
        order = {MM.STATUS_NEW: 0, MM.STATUS_UPDATED: 1}
        m_rows = sorted((r for r in rows if r["month_status"]),
                        key=lambda r: (order[r["month_status"]], r["no"]))
        _table(ws, m_headers, [w for _, _, w in MONTH_COLUMNS],
               [[r[f] for f in m_fields] for r in m_rows], st,
               link_col=m_fields.index("url"), header_row=MONTH_HEADER_ROW)

    # --- All Questions (the candidate sheet)
    ws = wb.create_sheet(MAIN_SHEET)
    headers = [h for h, _, _ in MAIN_COLUMNS]
    widths = [w for _, _, w in MAIN_COLUMNS]
    fields = [f for _, f, _ in MAIN_COLUMNS]
    _table(ws, headers, widths, [[r[f] for f in fields] for r in rows], st,
           link_col=fields.index("url"))

    # --- Projections (same model, different grouping; cheap)
    # one tab per examiner (presentation order), model (document) order inside each
    ex_headers = ["No.", "MIW Question ID", "Question", "Topic / Category",
                  "Question Bank", "MIW Answer Link"]
    for ex in meta["examiner_names"]:
        sub = [r for r in rows if ex in r["examiners"]]
        if not sub:
            continue
        ws = wb.create_sheet(ex)
        _table(ws, ex_headers, [6, 16, 80, 34, 14, 52],
               [[i, r["id"], r["question"], r["topic"], r["qb"], r["url"]]
                for i, r in enumerate(sub, 1)], st, link_col=5)

    ws = wb.create_sheet("By Question Bank")
    qbs, seen_files = [], []
    for r in rows:  # one row per QB page, in model order
        if r["file"] not in seen_files:
            seen_files.append(r["file"])
    for fname in seen_files:
        sub = [r for r in rows if r["file"] == fname]
        qbs.append([sub[0]["qb"], sub[0]["topic"], len(sub), sub[0]["url"].split("#")[0]])
    _table(ws, ["Question Bank", "Topic / Category", "Questions", "MIW Page Link"],
           [14, 44, 11, 52], qbs, st, link_col=3)

    if internal:
        ws = wb.create_sheet("Examiner Relationships")
        _table(ws, ["MIW Question ID", "Examiner(s)", "Relationship rows", "Question"],
               [16, 26, 16, 80],
               [[r["id"], r["examiners_text"], r["examiner_relationships"], r["question"]] for r in rows], st)
    return wb


# ------------------------------------------------------------------ output

def check_output_allowed(path):
    name = Path(path).name
    if name in PROTECTED_OUTPUTS or (FINAL_NAME_RE.search(name) and "WORKING" not in name):
        fail("refusing to write %s: historical release / final name is protected" % name)


def write_workbook(wb, path):
    check_output_allowed(path)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    os.replace(tmp, path)
    return path


def main(argv):
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--candidate-interim", action="store_true",
                    help="write the August 2026 interim candidate workbook")
    ap.add_argument("--candidate-share", action="store_true",
                    help="write the group-facing share workbook, with the month's "
                         "New & Updated sheet (see --month)")
    ap.add_argument("--month", default=DEFAULT_MONTH,
                    help="YYYY-MM for the New & Updated projection (default %s)" % DEFAULT_MONTH)
    ap.add_argument("--working-master", action="store_true",
                    help="also write the internal WORKING master (same model)")
    ap.add_argument("--out", help="explicit output path for the candidate workbook")
    ap.add_argument("--model-json", help="also dump the normalized export model here")
    ap.add_argument("--no-fresh-check", action="store_true", help=argparse.SUPPRESS)
    a = ap.parse_args(argv)
    if not (a.candidate_interim or a.candidate_share or a.out or a.model_json
            or a.working_master):
        ap.error("nothing to do: pass --candidate-share / --candidate-interim "
                 "(and/or --out/--model-json)")
    month = a.month if (a.candidate_share or a.out) else None

    try:
        meta, rows = build_export_model(check_fresh=not a.no_fresh_check, month=month)
        generated_at = _dt.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z")
        written = []
        if a.model_json:
            Path(a.model_json).write_text(json.dumps({"meta": meta, "rows": rows}, indent=1,
                                                     ensure_ascii=False) + "\n", encoding="utf-8")
            written.append(a.model_json)
        if a.candidate_share:
            name = SHARE_NAME_FMT % (meta["month_name"], meta["month_year"])
            written.append(str(write_workbook(
                render_workbook(meta, rows, generated_at), RELEASE_DIR / name)))
        if a.out:
            written.append(str(write_workbook(
                render_workbook(meta, rows, generated_at), Path(a.out))))
        if a.candidate_interim:
            # The interim workbook keeps its own (month-free) About banner, so it
            # is rendered from a model without the monthly projection.
            i_meta, i_rows = build_export_model(check_fresh=False)
            written.append(str(write_workbook(
                render_workbook(i_meta, i_rows, generated_at), RELEASE_DIR / INTERIM_NAME)))
        if a.working_master:
            w_meta, w_rows = build_export_model(check_fresh=False)
            written.append(str(write_workbook(
                render_workbook(w_meta, w_rows, generated_at, internal=True),
                RELEASE_DIR / WORKING_NAME)))
    except ExportFailure as e:
        print("EXPORT FAILURE: %s" % e)
        return 2
    print("export: %d canonical questions, %d files, %d examiner relationships / %d examiners"
          % (meta["canonical_questions"], meta["question_files"],
             meta["examiner_relationships"], meta["examiners"]))
    if month:
        print("        %s: %d NEW, %d UPDATED (%d rows) against baseline %s "
              "(%d questions)" % (meta["month_sheet"], meta["month_new"],
                                  meta["month_updated"], meta["month_rows"],
                                  meta["month_baseline_commit"],
                                  meta["month_baseline_questions"]))
    for w in written:
        print("wrote %s (%d bytes)" % (w, Path(w).stat().st_size))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
