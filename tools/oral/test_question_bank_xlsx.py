"""Focused tests for the question-bank XLSX exporter + validator.

  PYTHONIOENCODING=utf-8 python tools/oral/test_question_bank_xlsx.py

Proves (against a workbook exported from the CURRENT repo into scratch):
  one_row_per_question   1 canonical question -> exactly 1 main-sheet row, model
                         count == index count, examiner join is from the snapshot
  validator_passes       the freshly exported workbook validates clean (--interim)
  duplicate_id_refused   a duplicated id row is refused
  missing_row_refused    a deleted question row is refused
  phantom_id_refused     an id outside the corpus is refused
  wrong_link_refused     a link pointing at another question / a dead anchor /
                         a non-hyperlink cell / a malformed URL is refused
  text_drift_refused     an edited question cell is refused
  leak_refused           an internal token on the About sheet is refused
  dgma_refused           an "Official DGMA syllabus" header is refused
  interim_banner         the banner is present, and its removal is refused
  protected_outputs      the historical and final release names are never written
  determinism            two exports of the same model produce identical rows

No mutation framework: each case is a scratch copy + one edit + one validator
run, and the scratch directory is removed at the end.
"""
from __future__ import annotations

import copy
import io
import shutil
import subprocess
import sys
import tempfile
from contextlib import redirect_stdout
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import export_question_bank_xlsx as X  # noqa: E402
import validate_question_bank_xlsx as V  # noqa: E402

import openpyxl  # noqa: E402

RESULTS = []


def check(name, ok, detail=""):
    RESULTS.append((name, ok, detail))
    print("%s  %s%s" % ("PASS" if ok else "FAIL", name, ("  - " + detail) if detail else ""))


def validate(path, interim=True):
    buf = io.StringIO()
    with redirect_stdout(buf):
        rc = V.main([str(path)] + (["--interim"] if interim else []))
    return rc, buf.getvalue()


def refused(name, path, must_mention):
    rc, out = validate(path)
    check(name, rc != 0 and must_mention in out,
          "rc=%d, wanted %r in output; got: %s" % (rc, must_mention, out.strip().splitlines()[:2]))


def main():
    scratch = Path(tempfile.mkdtemp(prefix="qbxlsx_"))
    try:
        meta, rows = X.build_export_model(check_fresh=True)
        base = scratch / "MIW_test_INTERIM.xlsx"
        X.write_workbook(X.render_workbook(meta, rows, "test-time"), base)

        # one_row_per_question
        import json
        idx = json.loads(X.INDEX_PATH.read_text(encoding="utf-8"))
        snap = json.loads(X.SNAPSHOT_PATH.read_text(encoding="utf-8"))
        ws = openpyxl.load_workbook(base)[X.MAIN_SHEET]
        ids = [r[1] for r in ws.iter_rows(min_row=2, values_only=True) if r[1]]
        check("one_row_per_question",
              len(rows) == idx["total_questions"] == len(ids) == len(set(ids))
              and sum(r["examiner_relationships"] for r in rows) == snap["totals"]["relationships"],
              "model %d index %d sheet %d unique %d rel %d/%d" % (
                  len(rows), idx["total_questions"], len(ids), len(set(ids)),
                  sum(r["examiner_relationships"] for r in rows), snap["totals"]["relationships"]))

        rc, out = validate(base)
        check("validator_passes", rc == 0, out.strip())

        def mutated(name, edit):
            p = scratch / ("%s_INTERIM.xlsx" % name)
            shutil.copy(base, p)
            wb = openpyxl.load_workbook(p)
            edit(wb)
            wb.save(p)
            return p

        def dup(wb):
            s = wb[X.MAIN_SHEET]
            s.cell(row=3, column=2).value = s.cell(row=2, column=2).value
            s.cell(row=3, column=7).value = s.cell(row=2, column=7).value
            s.cell(row=3, column=7).hyperlink = s.cell(row=2, column=7).value
        refused("duplicate_id_refused", mutated("dup", dup), "duplicates")

        refused("missing_row_refused", mutated("miss", lambda wb: wb[X.MAIN_SHEET].delete_rows(5)), "missing")

        def phantom(wb):
            s = wb[X.MAIN_SHEET]
            s.cell(row=2, column=2).value = "QB99_Z#q1"
        refused("phantom_id_refused", mutated("phantom", phantom), "phantom")

        def wrong_link(wb):
            s = wb[X.MAIN_SHEET]
            s.cell(row=2, column=7).value = s.cell(row=3, column=7).value
            s.cell(row=2, column=7).hyperlink = s.cell(row=3, column=7).value
        refused("wrong_link_refused/other-question", mutated("wl1", wrong_link), "identity")

        def dead_anchor(wb):
            s = wb[X.MAIN_SHEET]
            qid = s.cell(row=2, column=2).value
            fname = qid.split("#")[0] + ".html"
            s.cell(row=2, column=2).value = fname[:-5] + "#q9999"
            url = "%s/meoclass1/%s#q9999" % (X.SITE_BASE, fname)
            s.cell(row=2, column=7).value = url
            s.cell(row=2, column=7).hyperlink = url
        # an id that is not canonical is phantom first; the anchor check is reached
        # through a valid id whose URL points at a dead anchor:
        def dead_anchor2(wb):
            s = wb[X.MAIN_SHEET]
            fname = s.cell(row=2, column=2).value.split("#")[0] + ".html"
            url = "%s/meoclass1/%s#q9999" % (X.SITE_BASE, fname)
            s.cell(row=2, column=7).value = url
            s.cell(row=2, column=7).hyperlink = url
        refused("wrong_link_refused/dead-anchor", mutated("wl2", dead_anchor2), "anchors")

        def not_hyperlink(wb):
            wb[X.MAIN_SHEET].cell(row=2, column=7).hyperlink = None
        refused("wrong_link_refused/not-a-hyperlink", mutated("wl3", not_hyperlink), "links")

        def malformed(wb):
            s = wb[X.MAIN_SHEET]
            s.cell(row=2, column=7).value = "http://example.com/QB1_A.html#q1"
            s.cell(row=2, column=7).hyperlink = "http://example.com/QB1_A.html#q1"
        refused("wrong_link_refused/malformed", mutated("wl4", malformed), "links")

        def text_drift(wb):
            c = wb[X.MAIN_SHEET].cell(row=2, column=3)
            c.value = c.value + " (edited)"
        refused("text_drift_refused", mutated("txt", text_drift), "text")

        def leak(wb):
            wb[X.ABOUT_SHEET].cell(row=20, column=2).value = "built from correction_lsavent_manifest.json"
        refused("leak_refused", mutated("leak", leak), "leakage")

        def dgma(wb):
            wb[X.MAIN_SHEET].cell(row=1, column=4).value = "Official DGMA syllabus"
        refused("dgma_refused", mutated("dgma", dgma), "dgma")

        about = openpyxl.load_workbook(base)[X.ABOUT_SHEET]
        banner = any("INTERIM" in str(c.value) for r in about.iter_rows() for c in r if c.value)
        check("interim_banner/present", banner)

        def strip_banner(wb):
            for r in wb[X.ABOUT_SHEET].iter_rows():
                for c in r:
                    if c.value and "INTERIM" in str(c.value):
                        c.value = str(c.value).replace("INTERIM", "")
        refused("interim_banner/removal-refused", mutated("nobanner", strip_banner), "interim")

        # protected_outputs: exporter refuses, and the real historical files are untouched
        before = {n: (X.RELEASE_DIR / n).stat().st_mtime_ns for n in X.PROTECTED_OUTPUTS
                  if (X.RELEASE_DIR / n).exists()}
        wb = X.render_workbook(meta, rows, "test-time")
        refusals = 0
        for n in sorted(X.PROTECTED_OUTPUTS) + ["MEO_QB_master_v27_final.xlsx", "MIW_FINAL.xlsx"]:
            try:
                X.write_workbook(wb, scratch / n)
            except X.ExportFailure:
                refusals += 1
        after = {n: (X.RELEASE_DIR / n).stat().st_mtime_ns for n in before}
        check("protected_outputs", refusals == len(X.PROTECTED_OUTPUTS) + 2 and before == after
              and not any((scratch / n).exists() for n in X.PROTECTED_OUTPUTS),
              "%d refusals, historical mtimes unchanged=%s" % (refusals, before == after))
        try:
            X.write_workbook(wb, scratch / X.WORKING_NAME)
            check("protected_outputs/working-allowed", (scratch / X.WORKING_NAME).exists())
        except X.ExportFailure as e:
            check("protected_outputs/working-allowed", False, str(e))

        # determinism: same model -> identical cell values (timestamp is an input)
        meta2, rows2 = X.build_export_model(check_fresh=False)
        check("determinism/model", rows == rows2 and meta == meta2)
        p2 = scratch / "second_INTERIM.xlsx"
        X.write_workbook(X.render_workbook(meta2, rows2, "test-time"), p2)
        a = [[c for c in r] for s in openpyxl.load_workbook(base).worksheets for r in s.iter_rows(values_only=True)]
        b = [[c for c in r] for s in openpyxl.load_workbook(p2).worksheets for r in s.iter_rows(values_only=True)]
        check("determinism/workbook", a == b)

        # syllabus fields exist in the model, empty, and never reach a sheet
        check("syllabus_fields_reserved_not_rendered",
              all(all(r[k] == "" for k in X.SYLLABUS_FIELDS) for r in rows)
              and not any(h in [c for _, c, _ in X.MAIN_COLUMNS] for h in X.SYLLABUS_FIELDS))

        # CLI smoke: --out into scratch
        r = subprocess.run([sys.executable, str(HERE / "export_question_bank_xlsx.py"),
                            "--out", str(scratch / "cli_INTERIM.xlsx")],
                           capture_output=True, text=True, encoding="utf-8",
                           env={**__import__("os").environ, "PYTHONIOENCODING": "utf-8"})
        check("cli_export", r.returncode == 0 and (scratch / "cli_INTERIM.xlsx").exists(),
              (r.stdout + r.stderr).strip()[-200:])
    finally:
        shutil.rmtree(scratch, ignore_errors=True)

    failed = [n for n, ok, _ in RESULTS if not ok]
    print("\n%d checks, %d failed" % (len(RESULTS), len(failed)))
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
