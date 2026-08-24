#!/usr/bin/env python3
"""Mutation suite for batch G2 - the fifth-submission fresh-intake production.

WHAT THIS SUITE ADDS OVER G1's
------------------------------
G1 proved the intake lane. G2 adds the surfaces the Founder asked to be kept
current in parallel with production, so the suite has to attack those too::

    remove an S005 occurrence          -> the newest evidence must not vanish
    edit a PRE-S005 occurrence         -> older evidence must be immutable
    file fresh evidence as historical  -> 788 must never absorb August
    name an examiner with no marker    -> attribution must stay earned
    duplicate a prior fresh ask        -> one ask, one card
    launder a follow-up into a card    -> the chain must not be flattened
    publish with no review record      -> review must be required
    un-ignore the raw carrier file     -> candidate names must not be committable
    drop a card from the master QB     -> the internal projection must be complete
    leak internal vocabulary into the
      candidate workbook               -> the shareable projection must stay clean
    create a FINAL workbook            -> the freeze gate must hold while open
    move the historical denominator    -> 788 must be pinned

RESTORATION IS IN A `finally`
-----------------------------
A previous session killed a mutation suite on a foreground timeout and it left
a deliberately-wrong content mutation live on a product page. Every mutation
here restores from a byte snapshot inside a `finally`, and the suite re-probes
after the run and reports residue explicitly.

  PYTHONIOENCODING=utf-8 python tools/oral/mutate_batch_g2.py
"""
from __future__ import annotations

import io
import json
import pathlib
import re
import shutil
import subprocess
import sys

HERE = pathlib.Path(__file__).resolve().parent
REPO = HERE.parent.parent
sys.path.insert(0, str(HERE))

from oral_bytes import enable_utf8_stdio      # noqa: E402

enable_utf8_stdio()

OUT = REPO / "meoclass1" / "oral-intelligence" / "examiner-audit"
RECORDS = OUT / "AUGUST2026_INTAKE_RECORDS.jsonl"
ADJ = OUT / "AUGUST2026_INTAKE_ADJUDICATIONS.json"
REVIEW = OUT / "AUGUST2026_BATCH_G2_REVIEW.json"
HIST = OUT / "ALL_SURVEYORS_SOURCE_RECORDS.jsonl"
MANIFEST = HERE / "batch_g2_manifest.json"
GITIGNORE = REPO / ".gitignore"

QBOOK = REPO / "docs" / "MIW-master-Question-bank"
MASTER = QBOOK / "MEO_QB_master_v27_WORKING.xlsx"
SHARE = QBOOK / "MIW_August2026_QuestionBank_INTERIM.xlsx"
FAKE_FINAL = QBOOK / "MIW_August2026_QuestionBank_v27_FINAL.xlsx"

INTAKE = "validate_oral_intake.py"
G2 = "validate_batch_g2.py"
INGEST = "ingest_august_intake.py"
XLSX = "validate_question_bank_xlsx.py"
TXT = ("docs/MIW-master-Question-bank/New questions from August orals/"
       "24 Aug 2026 oral questions.txt")


class Snapshot:
    """Byte snapshot restored by exact path. Never a git checkout: this suite
    runs against a dirty tree by definition, and `git checkout -- <file>` would
    discard the very edits under test."""

    def __init__(self, paths):
        self.data = {p: (p.read_bytes() if p.is_file() else None) for p in paths}

    def restore(self):
        bad = []
        for p, b in self.data.items():
            if b is None:
                if p.is_file():
                    p.unlink()
            else:
                p.write_bytes(b)
                if p.read_bytes() != b:
                    bad.append(str(p))
        return bad


def run(script, *args):
    r = subprocess.run([sys.executable, str(HERE / script), *args],
                       cwd=str(REPO), capture_output=True)
    return r.returncode, r.stdout.decode("utf-8", "replace")


def failing_checks(text):
    return {m.group(1) for m in re.finditer(r"^FAIL\s+(\S+)", text, re.M)}


def jsonl_rows(p):
    return [json.loads(l) for l in p.read_text(encoding="utf-8").splitlines() if l.strip()]


def write_jsonl(p, rows):
    p.write_text("\n".join(json.dumps(r, ensure_ascii=False) for r in rows) + "\n",
                 encoding="utf-8")


# ------------------------------------------------------------------ mutations
def m_remove_s005_occurrence():
    write_jsonl(RECORDS, [r for r in jsonl_rows(RECORDS) if r["occurrence_id"] != "AUG-0052"])


def m_edit_pre_s005_occurrence():
    rows = jsonl_rows(RECORDS)
    for r in rows:
        if r["occurrence_id"] == "AUG-0001":
            r["raw_question_text"] = r["raw_question_text"] + " (tidied)"
    write_jsonl(RECORDS, rows)


def m_fresh_into_historical():
    rows = jsonl_rows(HIST)
    rows.append({**rows[0], "source_id": "AUG-0052"})
    write_jsonl(HIST, rows)


def m_unevidenced_examiner():
    rows = jsonl_rows(RECORDS)
    for r in rows:
        if r["occurrence_id"] == "AUG-0041":
            r["examiner_attribution"] = "INDIVIDUALLY_ATTRIBUTED"
            r["attributed_examiner"] = "Rajappan"
            r["attribution_marker"] = None
    write_jsonl(RECORDS, rows)


def m_duplicate_prior_fresh_ask():
    d = json.loads(ADJ.read_text(encoding="utf-8"))
    src = next(a for a in d["adjudications"] if a["occurrence_id"] == "AUG-0035")
    for a in d["adjudications"]:
        if a["occurrence_id"] == "AUG-0050":
            a["classification"] = "GENUINE_NEW_QUESTION"
            a["ask"] = src["ask"]
            a.pop("prior_fresh_ask_ref", None)
            a["negative_search"] = src.get("negative_search", [])
    ADJ.write_text(json.dumps(d, indent=2, ensure_ascii=False), encoding="utf-8")


def m_launder_followup():
    d = json.loads(MANIFEST.read_text(encoding="utf-8"))
    d["cards"][1]["source_occurrence_ids"] = ["AUG-0039"]   # a FOLLOWUP occurrence
    MANIFEST.write_text(json.dumps(d, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def m_remove_review():
    REVIEW.unlink()


def m_unignore_raw_carrier():
    t = io.open(GITIGNORE, encoding="utf-8", newline="").read()
    io.open(GITIGNORE, "w", encoding="utf-8", newline="").write(
        t.replace("docs/MIW-master-Question-bank/**/*.txt",
                  "# docs/MIW-master-Question-bank/**/*.txt", 1))


def _xlsx(path):
    import openpyxl
    return openpyxl.load_workbook(path)


def m_master_drops_a_card():
    """Delete a row from the QUESTION sheet.

    worksheets[0] is the About sheet - fourteen rows of provenance. Deleting
    its last row changes nothing the validator looks at, so the first version
    of this mutation escaped while proving only that it had chosen the wrong
    sheet. The question sheet is found by name.
    """
    wb = _xlsx(MASTER)
    ws = next((w for w in wb.worksheets if w.max_row > 100), wb.worksheets[-1])
    ws.delete_rows(ws.max_row)
    wb.save(MASTER)


def m_shareable_leaks_private_evidence():
    wb = _xlsx(SHARE)
    ws = next((w for w in wb.worksheets if w.max_row > 100), wb.worksheets[-1])
    ws.cell(row=ws.max_row, column=1).value = "FUP-001 internal manifest digest"
    wb.save(SHARE)


def m_create_final_workbook():
    shutil.copyfile(SHARE, FAKE_FINAL)


def m_move_historical_denominator():
    write_jsonl(HIST, jsonl_rows(HIST)[:-1])


MUTATIONS = [
    ("A", "remove an S005 raw occurrence", [RECORDS], m_remove_s005_occurrence,
     (INTAKE,), "A2_every_intake_occurrence_adjudicated"),
    ("B", "edit a pre-S005 raw occurrence", [RECORDS], m_edit_pre_s005_occurrence,
     (INGEST,), "__ingest_drift__"),
    ("C", "file fresh August evidence in the historical ledger", [HIST],
     m_fresh_into_historical, (INTAKE,), "A4_no_intake_id_in_historical_ledger"),
    ("D", "attribute a question to the panel's examiner with no marker", [RECORDS],
     m_unevidenced_examiner, (INTAKE,), "A10_intake_attribution_matches_evidence"),
    ("E", "claim a second new card for a prior fresh ask", [ADJ],
     m_duplicate_prior_fresh_ask, (INTAKE,),
     "A12_no_two_occurrences_claim_the_same_new_card"),
    ("F", "launder an examiner follow-up into a new card", [MANIFEST],
     m_launder_followup, (G2,), "g2_action_kind_agrees_with_adjudication"),
    ("G", "publish the batch with no review record", [REVIEW], m_remove_review,
     (G2,), "g2_review_record_present"),
    ("H", "un-ignore the raw candidate carrier file", [GITIGNORE],
     m_unignore_raw_carrier, (INTAKE,), "P1_raw_carrier_files_are_git_ignored"),
    ("I", "drop a governed card from the master QB projection", [MASTER],
     m_master_drops_a_card, ("__master__",), "__xlsx__"),
    ("J", "leak internal vocabulary into the candidate QB projection", [SHARE],
     m_shareable_leaks_private_evidence, ("__share__",), "__xlsx__"),
    ("K", "create a FINAL August workbook while the intake window is open",
     [FAKE_FINAL], m_create_final_workbook, (INTAKE,),
     "Z1_no_final_august_workbook_while_intake_open"),
    ("L", "move the historical 788 denominator", [HIST],
     m_move_historical_denominator, (INTAKE,), "H1_historical_count_is_788"),
]


def probe(kind):
    if kind == INGEST:
        rc, _ = run(INGEST, "--txt", TXT, "--check")
        return rc, ({"__ingest_drift__"} if rc else set())
    if kind == "__master__":
        rc, _ = run(XLSX, str(MASTER))
        return rc, ({"__xlsx__"} if rc else set())
    if kind == "__share__":
        rc, _ = run(XLSX, str(SHARE), "--interim")
        return rc, ({"__xlsx__"} if rc else set())
    rc, out = run(kind)
    return rc, failing_checks(out)


ALL_PROBES = (INTAKE, G2, INGEST, "__master__", "__share__")


def main() -> int:
    if not MANIFEST.is_file():
        print("G2 manifest missing")
        return 2
    if FAKE_FINAL.is_file():
        print("stale freeze-gate fixture on disk: %s" % FAKE_FINAL.name)
        return 2

    print("--- preflight: every mutation must change bytes ---")
    no_ops = []
    for mid, desc, files, apply, _p, _c in MUTATIONS:
        snap = Snapshot(files)
        before = dict(snap.data)
        try:
            apply()
            after = {p: (p.read_bytes() if p.is_file() else None) for p in before}
            changed = any(before[p] != after[p] for p in before)
            delta = sum(len(after[p] or b"") - len(before[p] or b"") for p in before)
            print("%-3s %-58s %-7s byte_delta=%+d"
                  % (mid, desc, "applied" if changed else "NO-OP", delta))
            if not changed:
                no_ops.append(mid)
        except Exception as exc:                                   # noqa: BLE001
            print("%-3s ERROR %s: %s" % (mid, type(exc).__name__, exc))
            no_ops.append(mid)
        finally:
            bad = snap.restore()
        if bad:
            print("    RESTORE FAILED: %s" % bad)
            return 2
    if no_ops:
        print("\npreflight FAILED - no bytes changed: %s" % ", ".join(no_ops))
        return 1

    print("\n--- control: unmutated state ---")
    baseline = {}
    for k in ALL_PROBES:
        rc, failing = probe(k)
        print("    %-26s exit=%d failing=%s" % (k, rc, sorted(failing) or "none"))
        baseline[k] = failing
        if failing:
            print("PRE-RUN: %s already failing; a mutation caught here proves nothing." % k)
            return 2

    print("\n--- mutations ---")
    escapes, residue = [], []
    for mid, desc, files, apply, probes, expect in MUTATIONS:
        snap = Snapshot(files)
        try:
            apply()
            new = set()
            rc = 0
            for k in probes:
                rc, failing = probe(k)
                new |= (failing - baseline[k])
            caught = expect in new
            if not caught:
                escapes.append("%s (%s): expected %s, got %s"
                               % (mid, desc, expect, sorted(new) or "nothing"))
            print("%-3s %-58s %-8s exit=%d  %s"
                  % (mid, desc, "CAUGHT" if caught else "ESCAPED", rc,
                     expect if caught else (sorted(new) or "no failure")))
        finally:
            bad = snap.restore()
        if bad:
            residue.append("%s: %s" % (mid, bad))

    print("\n--- post-run: the tree must be green again ---")
    for k in ALL_PROBES:
        rc, failing = probe(k)
        print("    %-26s exit=%d failing=%s" % (k, rc, sorted(failing) or "none"))
        if failing:
            residue.append("%s still failing after restore: %s" % (k, sorted(failing)))

    print("\n%d mutations, %d escape(s), 0 no-op(s), %d residue"
          % (len(MUTATIONS), len(escapes), len(residue)))
    for e in escapes:
        print("  ESCAPE  %s" % e)
    for r in residue:
        print("  RESIDUE %s" % r)
    return 1 if (escapes or residue) else 0


if __name__ == "__main__":
    raise SystemExit(main())
