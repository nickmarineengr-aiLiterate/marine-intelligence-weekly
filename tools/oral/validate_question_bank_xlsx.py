"""Fast, export-specific validator for a generated MIW Oral question-bank workbook.

This is deliberately NOT the 50-gate Oral release. No product content changes
when a workbook is exported, so the only things to prove are that the workbook
is a faithful, complete, candidate-safe projection of the current canonical
corpus. Runs in seconds.

  PYTHONIOENCODING=utf-8 python tools/oral/validate_question_bank_xlsx.py <file.xlsx> [--interim]

Controls (every one must hold; exit 0 only then):
  sheets            About + All Questions present, About first, readable
  headers           main-sheet headers exactly as the exporter defines them
  count             main-sheet rows == canonical question count (rebuilt from repo)
  duplicates        duplicate canonical ids on the main sheet == 0
  missing           canonical ids absent from the main sheet == 0
  phantom           main-sheet ids absent from the canonical corpus == 0
  identity          each row's file/anchor (from its URL) matches its canonical id
  files             every linked file exists under meoclass1/
  anchors           every linked anchor is a q-card id on that live page
  text              question text == canonical display text, cell by cell
  topic/qb/examiner cells == model values (no renderer-side inference)
  links             every MIW Answer Link cell is a real hyperlink with the
                    expected https://marineintelligenceweekly.com/meoclass1/ shape
  usability         freeze panes, auto-filter, wrapped question column, sane widths,
                    no hidden rows/columns, no formulas
  leakage           no internal vocabulary on any sheet (digest, manifest, FUP-,
                    CORR-, validator, tier enum literals, filesystem paths, ...)
  dgma              no "Official DGMA" / "Annexure III" / syllabus-branch claims
  interim           (--interim) the About sheet carries the INTERIM banner and
                    the file name is not a protected/final release name

With --month YYYY-MM the share workbook's monthly view is certified too. The
month controls re-derive the projection from oral_monthly INDEPENDENTLY of the
exporter rather than reading the exporter's own answer back, so a bug in the
projection cannot certify itself:
  month/sheet       "<Month> <Year> - New & Updated" exists, is the SECOND sheet
                    and carries rows
  month/summary     the headline counts above the table equal the row counts
                    below it, and the total equals the canonical count
  month/status      every row is NEW or UPDATED, matching the independently
                    re-derived classification card for card
  month/provenance  every row traces to governed August evidence (a batch or
                    correction manifest action, or absence from / rewriting
                    since the pre-month baseline commit)
  month/complete    no projected card is missing from the sheet
  month/duplicates  no card appears on the month sheet twice
  month/rows        id, question, topic, qb, examiner and link on each month row
                    equal the canonical model, exactly as on the main sheet
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import export_question_bank_xlsx as X  # noqa: E402
import oral_monthly as MM  # noqa: E402

URL_RE = re.compile(r"^https://marineintelligenceweekly\.com/meoclass1/([A-Za-z0-9_]+\.html)#(q\d+)$")

LEAK_RE = re.compile(
    r"(sha256|digest|manifest|FUP-\d|CORR-[A-Z]|validator|mutation|confidence score|"
    r"tier_rank|REL-[A-Z]+-QB|\bce_tip\b|\binferred\b|\breported\b|[A-Za-z]:\\|/tools/oral/|"
    r"\.jsonl?\b|\.py\b|paywall|entitlement|razorpay|password|api[_ ]key)", re.I)
DGMA_RE = re.compile(r"official\s+dgma|annexure\s+iii|dgma\s+syllabus|official syllabus|circular\s*49", re.I)


def main(argv):
    if not argv or argv[0].startswith("--"):
        print(__doc__)
        return 2
    path = Path(argv[0])
    interim = "--interim" in argv
    month = next((a.split("=", 1)[1] for a in argv if a.startswith("--month=")), None)
    if month is None and "--month" in argv:
        i = argv.index("--month")
        month = argv[i + 1] if i + 1 < len(argv) else None
    problems = []

    def bad(ctrl, msg):
        problems.append("%-10s %s" % (ctrl, msg))

    import openpyxl
    try:
        wb = openpyxl.load_workbook(path)  # not read_only: need hyperlinks, dims, panes
    except Exception as e:  # noqa: BLE001
        print("FAIL  open    cannot open %s: %s" % (path, e))
        return 1

    try:
        meta, rows = X.build_export_model(check_fresh=True)
    except X.ExportFailure as e:
        print("FAIL  model   %s" % e)
        return 1
    by_id = {r["id"]: r for r in rows}
    canonical_text = {r["question"] for r in rows}

    # sheets
    names = wb.sheetnames
    if names[:1] != [X.ABOUT_SHEET]:
        bad("sheets", "first sheet is %r, expected %r" % (names[:1], X.ABOUT_SHEET))
    if X.MAIN_SHEET not in names:
        bad("sheets", "missing %r" % X.MAIN_SHEET)
        print("\n".join("FAIL  " + p for p in problems))
        return 1
    ws = wb[X.MAIN_SHEET]

    # headers
    headers = [c.value for c in ws[1]]
    expect = [h for h, _, _ in X.MAIN_COLUMNS]
    if headers[:len(expect)] != expect or any(headers[len(expect):]):
        bad("headers", "%r != %r" % (headers, expect))
    col = {h: i for i, h in enumerate(expect)}

    # rows
    data = [[c.value for c in r] for r in ws.iter_rows(min_row=2, max_col=len(expect))]
    data = [r for r in data if any(v not in (None, "") for v in r)]
    if len(data) != meta["canonical_questions"]:
        bad("count", "%d rows != %d canonical" % (len(data), meta["canonical_questions"]))
    ids = [r[col["MIW Question ID"]] for r in data]
    dup = {i for i in ids if ids.count(i) > 1}
    if dup:
        bad("duplicates", "%d duplicate ids e.g. %s" % (len(dup), sorted(dup)[:3]))
    missing = set(by_id) - set(ids)
    if missing:
        bad("missing", "%d canonical ids absent e.g. %s" % (len(missing), sorted(missing)[:3]))
    phantom = set(ids) - set(by_id)
    if phantom:
        bad("phantom", "%d ids not canonical e.g. %s" % (len(phantom), sorted(phantom)[:3]))

    # per-row identity / links / text / fields
    link_idx = col["MIW Answer Link"]
    for n, r in enumerate(data, 2):
        qid = r[col["MIW Question ID"]]
        m = by_id.get(qid)
        if not m:
            continue
        url = r[link_idx]
        um = URL_RE.match(str(url or ""))
        if not um:
            bad("links", "row %d: malformed url %r" % (n, url))
            continue
        fname, anchor = um.groups()
        if "%s#%s" % (Path(fname).stem, anchor) != qid:
            bad("identity", "row %d: url %s does not encode id %s" % (n, url, qid))
        if not (X.L.MEO / fname).is_file():
            bad("files", "row %d: %s missing" % (n, fname))
        elif not X.anchor_exists(fname, anchor):
            bad("anchors", "row %d: %s has no q-card %s" % (n, fname, anchor))
        cell = ws.cell(row=n, column=link_idx + 1)
        if not cell.hyperlink or cell.hyperlink.target != url:
            bad("links", "row %d: link cell is not a hyperlink to its own value" % n)
        if r[col["Question"]] != m["question"]:
            bad("text", "row %d (%s): text differs from canonical" % (n, qid))
        if r[col["Topic / Category"]] != m["topic"]:
            bad("topic", "row %d: topic %r != %r" % (n, r[col["Topic / Category"]], m["topic"]))
        if r[col["Question Bank"]] != m["qb"]:
            bad("qb", "row %d: qb %r != %r" % (n, r[col["Question Bank"]], m["qb"]))
        if (r[col["Examiner(s)"]] or "") != m["examiners_text"]:
            bad("examiner", "row %d: %r != %r" % (n, r[col["Examiner(s)"]], m["examiners_text"]))
        if r[col["No."]] != m["no"]:
            bad("order", "row %d: No. %r != %r" % (n, r[col["No."]], m["no"]))

    # --- monthly New & Updated view -------------------------------------
    if month:
        mname, myear = X.month_names(month)
        sheet = X.MONTH_SHEET_FMT % (mname, myear)
        if sheet not in names:
            bad("month/sheet", "missing %r" % sheet)
        elif names.index(sheet) != 1:
            bad("month/sheet", "%r is sheet %d; a candidate must meet it second"
                % (sheet, names.index(sheet) + 1))
        if sheet in names:
            # Independent re-derivation: never read the exporter's own answer back.
            try:
                proj = MM.classify(month, {r["id"]: r["question"] for r in rows})
            except MM.MonthlyFailure as e:
                bad("month/provenance", "cannot re-derive %s: %s" % (month, e))
                proj = None
            ms = wb[sheet]
            hrow = X.MONTH_HEADER_ROW
            m_expect = [h for h, _, _ in X.MONTH_COLUMNS]
            m_head = [c.value for c in ms[hrow]][:len(m_expect)]
            if m_head != m_expect:
                bad("month/sheet", "headers %r != %r" % (m_head, m_expect))
            mcol = {h: i for i, h in enumerate(m_expect)}
            mdata = [[c.value for c in r] for r in
                     ms.iter_rows(min_row=hrow + 1, max_col=len(m_expect))]
            mdata = [r for r in mdata if any(v not in (None, "") for v in r)]
            if not mdata:
                bad("month/sheet", "%r carries no rows" % sheet)
            mids = [r[mcol["MIW Question ID"]] for r in mdata]
            mdup = {i for i in mids if mids.count(i) > 1}
            if mdup:
                bad("month/duplicates", "%d card(s) listed twice e.g. %s"
                    % (len(mdup), sorted(mdup)[:3]))
            if proj:
                want = {q: MM.STATUS_NEW for q in proj["new"]}
                want.update({q: MM.STATUS_UPDATED for q in proj["updated"]})
                got = {r[mcol["MIW Question ID"]]: r[mcol["Status"]] for r in mdata}
                omitted = sorted(set(want) - set(got))
                if omitted:
                    bad("month/complete", "%d projected card(s) absent from the sheet: %s"
                        % (len(omitted), omitted[:5]))
                extra = sorted(set(got) - set(want))
                if extra:
                    bad("month/provenance", "%d sheet row(s) the projection does not "
                        "carry: %s" % (len(extra), extra[:5]))
                wrong = sorted(q for q in set(want) & set(got) if want[q] != got[q])
                if wrong:
                    bad("month/status", "%d card(s) classified against the evidence: %s"
                        % (len(wrong), [(q, got[q], want[q]) for q in wrong[:3]]))
                for q in set(want) & set(got):
                    if not proj["evidence"].get(q):
                        bad("month/provenance", "%s carries no governed evidence" % q)
                # headline counts above the table must agree with the table
                head = "\n".join(str(ms.cell(row=r, column=1).value or "")
                                 for r in range(1, hrow))
                for lbl, n in (("Questions in the bank", meta["canonical_questions"]),
                               ("New this month", len(proj["new"])),
                               ("Updated this month", len(proj["updated"]))):
                    if not re.search(r"%s:\s*%d\b" % (re.escape(lbl), n), head):
                        bad("month/summary", "headline does not state %s: %d" % (lbl, n))
                for st, n in ((MM.STATUS_NEW, len(proj["new"])),
                              (MM.STATUS_UPDATED, len(proj["updated"]))):
                    actual = sum(1 for r in mdata if r[mcol["Status"]] == st)
                    if actual != n:
                        bad("month/summary", "%d %s rows on the sheet, %d projected"
                            % (actual, st, n))
            # every month row is the canonical card, same contract as the main sheet
            mlink = mcol["MIW Answer Link"]
            for n, r in enumerate(mdata, hrow + 1):
                qid = r[mcol["MIW Question ID"]]
                m = by_id.get(qid)
                if not m:
                    bad("month/rows", "row %d: %r is not a canonical card" % (n, qid))
                    continue
                for header, field in (("Question", "question"), ("Topic / Category", "topic"),
                                      ("Question Bank", "qb"),
                                      ("Examiner(s)", "examiners_text")):
                    if (r[mcol[header]] or "") != m[field]:
                        bad("month/rows", "row %d (%s): %s differs from canonical"
                            % (n, qid, header))
                if r[mlink] != m["url"]:
                    bad("month/rows", "row %d (%s): link %r != %r" % (n, qid, r[mlink], m["url"]))
                elif not URL_RE.match(str(r[mlink] or "")):
                    bad("month/rows", "row %d: malformed url %r" % (n, r[mlink]))
                cell = ms.cell(row=n, column=mlink + 1)
                if not cell.hyperlink or cell.hyperlink.target != r[mlink]:
                    bad("month/rows", "row %d: link cell is not a hyperlink" % n)
            if ms.freeze_panes != "A%d" % (hrow + 1):
                bad("month/sheet", "freeze_panes %r, expected A%d"
                    % (ms.freeze_panes, hrow + 1))
            if not ms.auto_filter.ref:
                bad("month/sheet", "no auto-filter")
            if not ms.cell(row=hrow + 1, column=mcol["Question"] + 1).alignment.wrap_text:
                bad("month/sheet", "question column not wrapped")

    # usability
    if ws.freeze_panes != "A2":
        bad("usability", "main sheet freeze_panes %r" % ws.freeze_panes)
    if not ws.auto_filter.ref:
        bad("usability", "main sheet has no auto-filter")
    qcol = ws.cell(row=2, column=col["Question"] + 1)
    if not qcol.alignment.wrap_text:
        bad("usability", "question column not wrapped")
    from openpyxl.utils import get_column_letter
    for i, (_, _, w) in enumerate(X.MAIN_COLUMNS, 1):
        dim = ws.column_dimensions[get_column_letter(i)]
        if (dim.width or 0) < 5:
            bad("usability", "column %s width %r unusable" % (get_column_letter(i), dim.width))
    for s in wb.worksheets:
        if s.sheet_state != "visible":
            bad("usability", "sheet %r hidden" % s.title)
        for d in s.row_dimensions.values():
            if d.hidden:
                bad("usability", "%s: hidden row" % s.title)
        for d in s.column_dimensions.values():
            if d.hidden:
                bad("usability", "%s: hidden column" % s.title)

    # leakage / dgma / formulas / interim, across every sheet
    banner = False
    for s in wb.worksheets:
        for row in s.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                sv = str(v)
                if c.data_type == "f" or sv.startswith("="):
                    bad("formulas", "%s!%s has a formula" % (s.title, c.coordinate))
                if sv not in canonical_text:
                    # canonical question text may legitimately say e.g. "reported"
                    if LEAK_RE.search(sv):
                        bad("leakage", "%s!%s: %r" % (s.title, c.coordinate, sv[:60]))
                if DGMA_RE.search(sv):
                    bad("dgma", "%s!%s: %r" % (s.title, c.coordinate, sv[:60]))
                if "INTERIM" in sv and s.title == X.ABOUT_SHEET:
                    banner = True
    if interim:
        if not banner:
            bad("interim", "About sheet carries no INTERIM banner")
        if path.name in X.PROTECTED_OUTPUTS or X.FINAL_NAME_RE.search(path.name):
            bad("interim", "file name %s is a protected/final release name" % path.name)

    for p in problems:
        print("FAIL  " + p)
    if problems:
        print("%s: %d problem(s)" % (path.name, len(problems)))
        return 1
    print("PASS  %s: %d rows == %d canonical, 0 dup, 0 missing, 0 phantom, 0 dead files, "
          "0 dead anchors, %d hyperlinks, text/topic/qb/examiner exact, no leakage, no DGMA claims%s"
          % (path.name, len(data), meta["canonical_questions"], len(data),
             ", INTERIM banner present" if interim else ""))
    if month:
        p = MM.classify(month, {r["id"]: r["question"] for r in rows})
        print("PASS  %s: %d NEW + %d UPDATED = %d rows, 0 duplicate, 0 omitted, 0 "
              "unprovenanced; classification re-derived independently against "
              "baseline %s (%s, %d questions)"
              % (X.MONTH_SHEET_FMT % X.month_names(month), len(p["new"]), len(p["updated"]),
                 len(p["new"]) + len(p["updated"]), p["baseline_commit"],
                 p["baseline_date"], p["baseline_questions"]))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
