"""Audit qb_content_index.json and the two historical MIW workbooks against
the live Oral QB inventory.

The workbooks are external, git-ignored inputs; their paths are CLI arguments
so this tool stays portable across machines.

Usage:
  python tools/oral/audit_sources.py --master <MEO_QB_master_v26.xlsx> \
                                     --july   <MIW_July2026_QuestionBank_SHARE.xlsx>
Either workbook may be omitted; the JSON audit always runs.
"""
from __future__ import annotations

import argparse
import collections
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import oral_lib as L  # noqa: E402


# --------------------------------------------------------------------------
def audit_content_index(inv):
    path = L.MEO / "qb_content_index.json"
    d = json.loads(path.read_text(encoding="utf-8"))
    live_by_file = collections.defaultdict(dict)
    for r in inv:
        live_by_file[r["file"]][r["q_number"]] = r["question_text"]

    json_files = set(d["files"])
    live_files = set(live_by_file)
    json_q = sum(f.get("question_count") or 0 for f in d["files"].values())
    schema_gaps = sorted(
        k for k, f in d["files"].items()
        if "question_count" not in f or "questions" not in f
    )
    json_q_listed = sum(len(f.get("questions") or []) for f in d["files"].values())

    count_mismatch, text_mismatch, missing_q, extra_q = [], [], [], []
    for fname, blk in d["files"].items():
        live = live_by_file.get(fname)
        if live is None:
            continue
        if (blk.get("question_count") or 0) != len(live):
            count_mismatch.append(
                {"file": fname, "json": blk.get("question_count"), "html": len(live)}
            )
        jnums = {q["qnum"]: q["text"] for q in (blk.get("questions") or [])}
        for n, t in jnums.items():
            if n not in live:
                missing_q.append({"file": fname, "qnum": n, "json_text": t})
            elif L.norm(t) != L.norm(live[n]):
                sim = round(L.jaccard(t, live[n]), 3)
                text_mismatch.append(
                    {
                        "file": fname,
                        "qnum": n,
                        "similarity": sim,
                        "json_text": t,
                        "html_text": live[n],
                    }
                )
        for n in live:
            if n not in jnums:
                extra_q.append({"file": fname, "qnum": n, "html_text": live[n]})

    # duplicate question identifiers inside JSON
    dupes = []
    for fname, blk in d["files"].items():
        c = collections.Counter(q["qnum"] for q in (blk.get("questions") or []))
        for n, k in c.items():
            if k > 1:
                dupes.append({"file": fname, "qnum": n, "occurrences": k})

    severe = len(missing_q) + len(count_mismatch) + len(dupes)
    drifted = sum(1 for m in text_mismatch if m["similarity"] < 0.7)
    if json_files - live_files or severe:
        klass = "DERIVED_BUT_STALE"
    elif drifted:
        klass = "DERIVED_BUT_STALE"
    else:
        klass = "DERIVED_AND_CURRENT"
    return {
        "path": str(path.relative_to(L.REPO)).replace("\\", "/"),
        "manifest_version": d["manifest_version"],
        "generated": d["generated"],
        "generated_by": d["generated_by"],
        "headline_total_questions": d["total_questions"],
        "headline_total_files": d["total_files"],
        "json_files_declared": len(json_files),
        "json_question_count_sum": json_q,
        "json_questions_listed": json_q_listed,
        "html_files": len(live_files),
        "html_questions": len(inv),
        "files_in_json_absent_from_repo": sorted(json_files - live_files),
        "live_qb_files_absent_from_json": sorted(live_files - json_files),
        "file_question_count_mismatches": count_mismatch,
        "questions_in_json_absent_from_html": missing_q,
        "questions_in_html_absent_from_json": extra_q,
        "question_text_mismatches": len(text_mismatch),
        "question_text_mismatches_below_0_7": drifted,
        "text_mismatch_examples": sorted(text_mismatch, key=lambda m: m["similarity"])[:15],
        "duplicate_question_identifiers": dupes,
        "entries_missing_schema_fields": schema_gaps,
        "classification": klass,
    }


# --------------------------------------------------------------------------
def _rows(ws, header_row):
    hdr = None
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < header_row:
            continue
        if hdr is None:
            hdr = [str(c).strip() if c is not None else "" for c in row]
            continue
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        out.append({hdr[j]: row[j] for j in range(min(len(hdr), len(row)))})
    return hdr, out


def audit_master(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    tracker = []
    examiner_sheets = {}
    for ws in wb.worksheets:
        hr = 2 if ws.title in {"Senthil", "Srivastava", "Rajappan", "Simon", "Nair", "Paul"} else 0
        hdr, rows = _rows(ws, hr)
        sheets.append(
            {
                "sheet": ws.title,
                "header_row_index": hr,
                "columns": hdr,
                "data_rows": len(rows),
                "declared_in_title": (ws.cell(1, 1).value if hr else None),
            }
        )
        if ws.title == "All Questions":
            tracker = rows
        elif hr == 2:
            examiner_sheets[ws.title] = rows
    wb.close()
    return {"sheets": sheets, "tracker": tracker, "examiner_sheets": examiner_sheets}


def audit_july(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    per_examiner = {}
    all_q = []
    new_q = []
    EX = {"Nair", "Simon", "Rajappan", "Srivastava", "Senthil", "Paul"}
    for ws in wb.worksheets:
        hdr, rows = _rows(ws, 0)
        # capture hyperlink target per row (column C for examiner sheets)
        links = {}
        for row in ws.iter_rows():
            for c in row:
                if c.hyperlink is not None:
                    links[c.row] = c.hyperlink.target
        sheets.append(
            {
                "sheet": ws.title,
                "columns": hdr,
                "data_rows": len(rows),
                "hyperlinks": len(links),
                "hyperlink_has_anchor": sum(1 for t in links.values() if t and "#" in t),
            }
        )
        if ws.title in EX:
            out = []
            for i, r in enumerate(rows):
                out.append(
                    {
                        "no": r.get("No."),
                        "question": r.get("Question"),
                        "link": links.get(i + 2),
                    }
                )
            per_examiner[ws.title] = out
        elif ws.title == "All Questions":
            for i, r in enumerate(rows):
                r["link"] = links.get(i + 2)
                all_q.append(r)
        elif ws.title.startswith("New Questions"):
            for i, r in enumerate(rows):
                r["link"] = links.get(i + 2)
                new_q.append(r)
    wb.close()
    return {
        "sheets": sheets,
        "per_examiner": per_examiner,
        "all_questions": all_q,
        "new_questions": new_q,
    }


# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--master")
    ap.add_argument("--july")
    a = ap.parse_args()

    inv = L.build_inventory()
    out = {"content_index": audit_content_index(inv)}

    if a.master:
        m = audit_master(Path(a.master))
        L.jdump(m["tracker"], "MASTER_TRACKER_ROWS.json")
        L.jdump(m["examiner_sheets"], "MASTER_EXAMINER_SHEETS.json")
        tr = m["tracker"]
        out["master_workbook"] = {
            "file": Path(a.master).name,
            "sheets": m["sheets"],
            "tracker_rows": len(tr),
            "tracker_examiners": dict(
                collections.Counter(str(r.get("Examiner") or "").strip() for r in tr)
            ),
            "build_status_values": dict(
                collections.Counter(str(r.get("Build_Status") or "").strip()[:60] for r in tr)
            ),
            "match_confidence_values": dict(
                collections.Counter(str(r.get("Match_Confidence") or "").strip()[:40] for r in tr)
            ),
            "rows_with_live_file": sum(1 for r in tr if r.get("Live_File")),
            "rows_with_anchor_in_live_file": sum(
                1 for r in tr if r.get("Live_File") and "#" in str(r.get("Live_File"))
            ),
            "examiner_sheet_rows": {k: len(v) for k, v in m["examiner_sheets"].items()},
        }

    if a.july:
        j = audit_july(Path(a.july))
        L.jdump(j["per_examiner"], "JULY_EXAMINER_SHEETS.json")
        L.jdump(j["new_questions"], "JULY_NEW_QUESTIONS.json")
        out["july_workbook"] = {
            "file": Path(a.july).name,
            "sheets": j["sheets"],
            "per_examiner_rows": {k: len(v) for k, v in j["per_examiner"].items()},
            "per_examiner_total": sum(len(v) for v in j["per_examiner"].values()),
            "all_questions_rows": len(j["all_questions"]),
            "new_questions_rows": len(j["new_questions"]),
            "new_questions_examiners": dict(
                collections.Counter(
                    str(r.get("Examiner") or "").strip() for r in j["new_questions"]
                )
            ),
        }

    L.jdump(out, "SOURCE_AUDIT.json")
    print(json.dumps(out, indent=2, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
